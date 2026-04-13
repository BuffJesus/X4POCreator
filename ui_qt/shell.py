"""Main application shell — sidebar navigation + stacked content pages.

Ported from the Tuner's ``cpp/app/main.cpp`` MainWindow pattern:

    [ sidebar ][      stacked content page      ]
    [ Load    ][                                 ]
    [ Filter  ][    current page fills available]
    [ Bulk    ][    space                        ]
    [ Review  ][                                 ]
    [ Help    ][                                 ]
    [---------][                                 ]
    [ v0.10.0 ][                                 ]

The sidebar is a ``QListWidget`` styled via ``theme_qt.sidebar_style()`` so
the 3px ``accent_primary`` left bar on the selected item matches the same
"attention here" grammar used by ``card_style(accent=...)`` on content
cards.

Pages are placeholder ``QWidget`` instances in alpha1.  Each alpha phase
replaces one placeholder with the real surface:

    alpha1  (this)  — empty shell, all placeholders
    alpha2          — Load + Help
    alpha3          — Bulk grid
    alpha4          — Review, Export, dialogs
    beta1           — Command palette, shortcuts, polish
    release         — Delete tk, rename qt → primary

This module does not import any flow modules directly — pages hook into
them themselves.  Keep the shell UI-only so it's trivial to test.
"""

from __future__ import annotations

import os

from PySide6.QtCore import Qt, QSize, QThread, QEvent
from PySide6.QtGui import QAction, QKeySequence
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QStackedWidget,
    QStatusBar,
    QVBoxLayout,
    QWidget,
)

import export_flow
import maintenance
import maintenance_flow
import theme as t
import theme_qt as tq
from debug_log import write_debug
from ui_qt.load_tab import LoadTab
from ui_qt.help_tab import HelpTab
from ui_qt.filter_tab import FilterTab
from ui_qt.bulk_tab import BulkTab
from ui_qt.review_tab import ReviewTab
from ui_qt.session_controller import QtSessionController
from ui_qt.undo_stack import BulkUndoStack


# Sidebar nav items — matches the confirmed tkinter tab order for v0.10.0
# alpha1: Load → Filter → Bulk → Review → Help.  The filter tab in tk is a
# pair (line-code exclusion + customer exclusion); for now the Qt port
# keeps them as a single "Filter" entry — we'll decide whether to split or
# combine when alpha2 lands.
NAV_ITEMS = [
    ("Load",   "Load CSV source reports"),
    ("Filter", "Exclude line codes and customers"),
    ("Bulk",   "Assign vendors and review draft quantities"),
    ("Review", "Review exceptions and export POs"),
    ("Help",   "Documentation, shortcuts, and release notes"),
]


def _placeholder_page(title: str, breadcrumb: str, phase: str) -> QWidget:
    """Build a placeholder page for a surface not yet ported.

    Each placeholder shows the tab's title + breadcrumb banner (using the
    shared ``tab_header_html`` helper so every page's header reads the
    same way) and a dim message explaining which alpha will port it.
    """
    page = QWidget()
    layout = QVBoxLayout(page)
    layout.setContentsMargins(t.SPACE_LG, t.SPACE_LG, t.SPACE_LG, t.SPACE_LG)
    layout.setSpacing(t.SPACE_MD)

    header = QLabel()
    header.setTextFormat(Qt.RichText)
    header.setText(tq.tab_header_html(title, breadcrumb))
    header.setStyleSheet(tq.tab_header_style())
    layout.addWidget(header)

    body = QLabel(f"This surface is migrating in {phase}.\n\n"
                  f"Until then, use the tkinter build (POBuilder.exe) for this workflow.")
    body.setAlignment(Qt.AlignCenter)
    body.setWordWrap(True)
    body.setStyleSheet(
        f"color: {t.TEXT_DIM}; "
        f"font-size: {t.FONT_BODY}px; "
        f"background-color: {t.BG_PANEL}; "
        f"border: 1px solid {t.BORDER_SOFT}; "
        f"border-radius: {t.RADIUS_MD}px; "
        f"padding: {t.SPACE_XL}px;"
    )
    layout.addWidget(body, stretch=1)

    return page


# Custom events for thread→GUI communication in the update flow.
_UPDATE_READY_TYPE = QEvent.Type(QEvent.registerEventType())
_UPDATE_FAILED_TYPE = QEvent.Type(QEvent.registerEventType())


class _UpdateReadyEvent(QEvent):
    def __init__(self, staging: str, current_exe: str):
        super().__init__(_UPDATE_READY_TYPE)
        self.staging = staging
        self.current_exe = current_exe


class _UpdateFailedEvent(QEvent):
    def __init__(self, error: str):
        super().__init__(_UPDATE_FAILED_TYPE)
        self.error = error


class POBuilderShell(QMainWindow):
    """Top-level window: sidebar + stacked content + status bar."""

    def __init__(self, app_version: str = "", *, app_settings: dict | None = None, parent=None):
        super().__init__(parent)
        self.app_version = app_version
        self.app_settings = app_settings if app_settings is not None else {}
        self._loaded_report_paths: dict[str, str] = {}
        self._vendor_export_scope_overrides: dict[str, str] = {}
        self._loading_depth = 0
        self._loading_message = ""
        self.setWindowTitle(f"PO Builder (Qt) — v{app_version}" if app_version else "PO Builder (Qt)")
        self.resize(1280, 800)
        self.setMinimumSize(QSize(960, 600))

        # Session controller — owns business state and runs the pipeline.
        self.controller = QtSessionController(app_settings=self.app_settings)
        self.undo_stack = BulkUndoStack()
        try:
            self.controller.load_persistent_state()
        except Exception:
            pass  # Non-fatal — order_rules etc. may not exist yet

        # Populated during page construction.
        self.load_tab: LoadTab | None = None
        self.filter_tab: FilterTab | None = None
        self.bulk_tab: BulkTab | None = None
        self.review_tab: ReviewTab | None = None

        central = QWidget()
        h = QHBoxLayout(central)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(0)

        # ─── Sidebar ──────────────────────────────────────────────────
        self.sidebar = QListWidget()
        self.sidebar.setFixedWidth(170)
        self.sidebar.setStyleSheet(tq.sidebar_style())
        self.sidebar.setFocusPolicy(Qt.NoFocus)
        self._workflow_state: dict[str, str] = {}  # label → "pending"|"done"|"active"
        for label, _tooltip in NAV_ITEMS:
            item = QListWidgetItem(label)
            item.setSizeHint(QSize(170, 44))
            self.sidebar.addItem(item)
        h.addWidget(self.sidebar)

        # ─── Stacked content ──────────────────────────────────────────
        self.stack = QStackedWidget()
        self.stack.setStyleSheet(f"background-color: {t.BG_BASE};")
        # Each nav entry maps to either a real alpha2+ surface or a
        # placeholder page that explains which phase will port it.
        for label, tooltip in NAV_ITEMS:
            if label == "Load":
                self.load_tab = LoadTab(app_settings=self.app_settings)
                self.load_tab.load_finished.connect(self._on_load_finished)
                self.load_tab.load_failed.connect(self._on_load_failed)
                self.stack.addWidget(self.load_tab)
            elif label == "Bulk":
                self.bulk_tab = BulkTab()
                self.bulk_tab.vendor_applied.connect(self._on_vendor_applied)
                self.bulk_tab.rows_removed.connect(self._on_rows_removed)
                self.bulk_tab.remove_not_needed.connect(self._on_remove_not_needed)
                self.bulk_tab.ignore_item.connect(self._on_ignore_item)
                self.bulk_tab.edit_committed.connect(self._on_edit_committed)
                self.bulk_tab.cycle_changed.connect(self._on_cycle_changed)
                self.bulk_tab.draft_review_requested.connect(self._on_draft_review)
                self.bulk_tab.undo_requested.connect(self._on_undo)
                self.bulk_tab.redo_requested.connect(self._on_redo)
                # Workflow dialog signals
                self.bulk_tab.vendor_review_requested.connect(self._on_vendor_review)
                self.bulk_tab.session_diff_requested.connect(self._on_session_diff)
                self.bulk_tab.supplier_map_requested.connect(self._on_supplier_map)
                self.bulk_tab.qoh_review_requested.connect(self._on_qoh_review)
                self.bulk_tab.skip_actions_requested.connect(self._on_skip_actions)
                self.bulk_tab.session_history_requested.connect(self._on_session_history)
                self.bulk_tab.ignored_items_requested.connect(self._on_ignored_items)
                self.bulk_tab.vendor_manager_requested.connect(self._on_vendor_manager)
                self.bulk_tab.settings_requested.connect(self._on_settings)
                self.bulk_tab.export_dead_stock.connect(self._on_export_dead_stock)
                self.bulk_tab.export_deferred.connect(self._on_export_deferred)
                self.bulk_tab.export_session_summary.connect(self._on_export_session_summary)
                self.bulk_tab.model.edit_callback = self._on_cell_edit
                self.bulk_tab.model.before_edit_callback = self._on_before_cell_edit
                self.stack.addWidget(self.bulk_tab)
            elif label == "Help":
                self.stack.addWidget(HelpTab())
            elif label == "Filter":
                self.filter_tab = FilterTab()
                self.filter_tab.filters_applied.connect(self._on_filters_applied)
                self.stack.addWidget(self.filter_tab)
            elif label == "Review":
                self.review_tab = ReviewTab()
                self.review_tab.export_requested.connect(self._on_export_requested)
                self.review_tab.items_changed.connect(self._on_review_items_changed)
                self.stack.addWidget(self.review_tab)
        h.addWidget(self.stack, stretch=1)

        self.setCentralWidget(central)

        # Navigate on sidebar selection — refresh review tab when switching to it
        self.sidebar.currentRowChanged.connect(self._on_nav_changed)
        self.sidebar.setCurrentRow(0)

        # ─── Status bar ───────────────────────────────────────────────
        self.status = QStatusBar()
        self.status.setSizeGripEnabled(False)
        self._status_item_count = QLabel("")
        self._status_item_count.setStyleSheet(
            f"color: {t.TEXT_MUTED}; padding: 0 {t.SPACE_MD}px; font-size: {t.FONT_SMALL}px;"
        )
        self.status.addPermanentWidget(self._status_item_count)
        self._status_data_dir = QLabel("")
        self._status_data_dir.setStyleSheet(
            f"color: {t.TEXT_DIM}; padding: 0 {t.SPACE_SM}px; font-size: {t.FONT_MICRO}px;"
        )
        self.status.addPermanentWidget(self._status_data_dir)
        version_label = QLabel(f"v{app_version} — Qt" if app_version else "Qt")
        version_label.setStyleSheet(f"color: {t.TEXT_DIM}; padding: 0 {t.SPACE_MD}px;")
        self.status.addPermanentWidget(version_label)
        self.setStatusBar(self.status)
        # Show data directory
        self._status_data_dir.setText(self.controller.data_dir)

        # Ctrl+K placeholder — wired in beta1 with the Qt command palette.
        # Register the shortcut now so the keystroke is a known no-op
        # rather than a random keypress dropping through to whatever has
        # focus.
        self._ctrl_k_action = QAction("Command Palette", self)
        self._ctrl_k_action.setShortcut(QKeySequence("Ctrl+K"))
        self._ctrl_k_action.triggered.connect(self._on_ctrl_k)
        self.addAction(self._ctrl_k_action)

        self._shortcut_action = QAction("Keyboard Shortcuts", self)
        self._shortcut_action.setShortcut(QKeySequence("?"))
        self._shortcut_action.triggered.connect(self._on_show_shortcuts)
        self.addAction(self._shortcut_action)

        # Background update check — deferred until after the first load
        # completes so it never competes with the parse worker for the GIL.
        self._update_release: dict | None = None
        self._version_label = version_label
        self._update_check_done = False

    def _mark_step(self, label: str, state: str):
        """Update a sidebar item's workflow state indicator.

        state: "pending" (dim dot), "done" (green check), "active" (blue arrow)
        """
        self._workflow_state[label] = state
        _ICONS = {"done": "\u2713 ", "active": "\u25B8 ", "pending": ""}
        for i, (nav_label, _) in enumerate(NAV_ITEMS):
            icon = _ICONS.get(self._workflow_state.get(nav_label, "pending"), "")
            item = self.sidebar.item(i)
            if item:
                item.setText(f"{icon}{nav_label}")

    def _on_nav_changed(self, row: int):
        self.stack.setCurrentIndex(row)
        # Refresh review tab when navigating to it
        label = NAV_ITEMS[row][0] if 0 <= row < len(NAV_ITEMS) else ""
        if label == "Review" and self.review_tab and self.bulk_tab:
            assigned = [
                item for item in self.bulk_tab.model.items
                if str(item.get("vendor", "")).strip()
            ]
            self.review_tab.set_items(assigned)

    def _on_review_items_changed(self):
        """Sync bulk model after edits or removals in the review tab."""
        if self.bulk_tab:
            self.bulk_tab.model.refresh_all()
            self.bulk_tab._update_summary()

    def _on_ctrl_k(self):
        from ui_qt.command_palette import CommandPaletteDialog
        dialog = CommandPaletteDialog(self, parent=self)
        dialog.exec()

    def _on_show_shortcuts(self):
        from ui_qt.shortcut_overlay import ShortcutOverlayDialog
        dialog = ShortcutOverlayDialog(parent=self)
        dialog.exec()

    # ── Update check ─────────────────────────────────────────────

    def _start_update_check(self):
        """Run a background check for newer GitHub releases."""
        check_enabled = self.app_settings.get("check_for_updates_on_startup", True)
        if not check_enabled:
            return

        class _Worker(QThread):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.release = None

            def run(self):
                import update_check
                self.release = update_check.check_for_update()

        self._update_worker = _Worker(self)
        self._update_worker.finished.connect(self._on_update_check_done)
        self._update_worker.start()

    def _on_update_check_done(self):
        release = getattr(self._update_worker, "release", None)
        if not release:
            return
        self._update_release = release
        tag = release.get("tag_name", "")
        self._version_label.setText(f"Update available: {tag}")
        self._version_label.setStyleSheet(
            f"color: {t.ACCENT_PRIMARY}; padding: 0 {t.SPACE_MD}px; "
            f"font-weight: bold; cursor: pointer;"
        )
        self._version_label.mousePressEvent = lambda _: self._show_update_dialog()
        self.status.showMessage(f"Update available: {tag}", 10000)

    def _show_update_dialog(self):
        import webbrowser
        import update_flow
        release = self._update_release
        if not release:
            return
        tag = release.get("tag_name", "")
        name = release.get("name") or tag
        published = release.get("published_at", "")[:10]
        exe_url = update_flow.find_exe_asset(release)
        can_auto = update_flow.can_self_update() and exe_url

        if can_auto:
            msg = (
                f"A newer release is available.\n\n"
                f"Current: v{self.app_version}\n"
                f"Latest: {name} ({tag})\n"
                f"Published: {published}\n\n"
                f"Download and install now? The app will restart."
            )
            answer = QMessageBox.question(
                self, "Update Available", msg,
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            )
            if answer == QMessageBox.Yes:
                self._do_auto_update(exe_url)
            elif answer == QMessageBox.No:
                url = release.get("html_url", "")
                if url:
                    webbrowser.open(url)
        else:
            msg = (
                f"A newer release is available.\n\n"
                f"Current: v{self.app_version}\n"
                f"Latest: {name} ({tag})\n"
                f"Published: {published}\n\n"
                f"Open the release page to download?"
            )
            if QMessageBox.question(self, "Update Available", msg) == QMessageBox.Yes:
                url = release.get("html_url", "")
                if url:
                    webbrowser.open(url)

    def _do_auto_update(self, exe_url: str):
        import sys
        import threading
        import update_flow
        current_exe = sys.executable
        staging = update_flow.staging_path_for(current_exe)
        self.status.showMessage("Downloading update...")

        def _worker():
            try:
                update_flow.download_update(exe_url, staging)
            except Exception as exc:
                update_flow.cleanup_staging(staging)
                QApplication.instance().postEvent(self, _UpdateFailedEvent(str(exc)))
                return
            QApplication.instance().postEvent(self, _UpdateReadyEvent(staging, current_exe))

        threading.Thread(target=_worker, daemon=True).start()

    def customEvent(self, event):
        if isinstance(event, _UpdateReadyEvent):
            self._on_update_downloaded(event.staging, event.current_exe)
        elif isinstance(event, _UpdateFailedEvent):
            QMessageBox.critical(self, "Update Failed",
                                 f"Download failed:\n{event.error}\n\nPlease download manually.")
            self.status.clearMessage()

    def _on_update_downloaded(self, staging: str, current_exe: str):
        import update_flow
        try:
            script = update_flow.write_updater_script(staging, current_exe)
        except Exception as exc:
            update_flow.cleanup_staging(staging)
            QMessageBox.critical(self, "Update Failed", f"Could not prepare update:\n{exc}")
            self.status.clearMessage()
            return
        answer = QMessageBox.question(
            self, "Ready to Install",
            "Update downloaded. Restart now to install?",
        )
        if answer == QMessageBox.Yes:
            update_flow.launch_updater_and_exit(self, script)

    def _on_load_finished(self, result: dict):
        """Handle a completed parse result from the Load tab."""
        if self.load_tab is not None:
            self._loaded_report_paths = dict(self.load_tab.current_paths())
        sales_items = result.get("sales_items") or []
        warnings = result.get("warnings") or []
        startup_rows = result.get("startup_warning_rows") or []
        write_debug("qt.load_finished",
                     sales_items=len(sales_items),
                     warnings=len(warnings),
                     startup_rows=len(startup_rows))

        self.status.showMessage(
            f"Loaded {len(sales_items)} sales item(s) — "
            f"{len(warnings)} warning(s), {len(startup_rows)} startup row(s)",
            10000,
        )

        if not sales_items:
            QMessageBox.warning(
                self,
                "No Data",
                "No items found in the loaded source files.  Check the file format.",
            )
            return

        # Apply load result to session
        ctrl = self.controller
        ctrl.apply_load_result(result)

        # Populate filter tab with line codes and customers
        if self.filter_tab is not None:
            from collections import Counter, defaultdict
            lc_counts = defaultdict(int)
            for item in ctrl.session.sales_items:
                lc_counts[item["line_code"]] += 1
            line_codes = sorted(ctrl.session.all_line_codes)

            cust_counts = Counter()
            cust_names = {}
            for si in ctrl.session.suspended_items:
                code = si.get("customer_code", "")
                if code:
                    cust_counts[code] += 1
                    cust_names.setdefault(code, si.get("customer", ""))
            customers = [
                (code, cust_names.get(code, ""), cust_counts[code])
                for code in sorted(cust_counts.keys())
            ]

            self.filter_tab.populate(line_codes, dict(lc_counts), customers)

            self._mark_step("Load", "done")
            self._mark_step("Filter", "active")

            # Switch to Filter tab
            filter_idx = next(
                (i for i, (label, _) in enumerate(NAV_ITEMS) if label == "Filter"),
                None,
            )
            if filter_idx is not None:
                self.sidebar.setCurrentRow(filter_idx)

        item_count = len(sales_items)
        self.status.showMessage(
            f"Loaded {item_count:,} items — adjust filters then click Continue", 15000,
        )

        # Deferred update check — runs after first load so it never
        # competes with the parse worker for the GIL.
        if not self._update_check_done:
            self._update_check_done = True
            self._start_update_check()

        if warnings:
            summary_lines = ["Top warnings:"]
            for title, _body in warnings[:5]:
                summary_lines.append(f"  \u2022 {title}")
            if len(warnings) > 5:
                summary_lines.append(f"  \u2026 and {len(warnings) - 5} more")
            QMessageBox.information(
                self, "Load Complete",
                f"Loaded {item_count} items.\n\n" + "\n".join(summary_lines),
            )

    def _on_load_failed(self, error: str):
        write_debug("qt.load_failed", error=error)
        self.status.showMessage(f"Load failed: {error}", 15000)

    def _on_filters_applied(self, excluded_lc: set, excluded_cust: set):
        """Run the assignment pipeline with the selected exclusions (background thread)."""
        from ui_qt.assignment_worker import AssignmentWorker
        write_debug("qt.filters_applied",
                     excluded_lc=len(excluded_lc),
                     excluded_cust=len(excluded_cust),
                     lc_sample=",".join(sorted(excluded_lc)[:5]))
        ctrl = self.controller
        ctrl.excluded_line_codes = set(excluded_lc)
        ctrl.excluded_customers = set(excluded_cust)
        self._excluded_lc_snapshot = set(excluded_lc)
        self._excluded_cust_snapshot = set(excluded_cust)

        # Disable sidebar and filter button during pipeline
        self.sidebar.setEnabled(False)
        if self.filter_tab:
            self.filter_tab._apply_btn.setEnabled(False)
        self.status.showMessage("Preparing assignment session\u2026")

        # Run pipeline on a background thread
        self._assign_thread = QThread(self)
        self._assign_worker = AssignmentWorker(ctrl)
        self._assign_worker.moveToThread(self._assign_thread)
        self._assign_thread.started.connect(self._assign_worker.run)
        self._assign_worker.progress.connect(
            lambda msg: self.status.showMessage(msg),
            Qt.QueuedConnection,
        )
        self._assign_worker.finished.connect(self._on_assignment_finished, Qt.QueuedConnection)
        self._assign_worker.failed.connect(self._on_assignment_failed, Qt.QueuedConnection)
        self._assign_worker.finished.connect(self._assign_thread.quit)
        self._assign_worker.failed.connect(self._assign_thread.quit)
        self._assign_thread.finished.connect(self._cleanup_assign_thread)
        self._assign_thread.start()

    def _cleanup_assign_thread(self):
        if hasattr(self, "_assign_worker") and self._assign_worker:
            self._assign_worker.deleteLater()
            self._assign_worker = None
        if hasattr(self, "_assign_thread") and self._assign_thread:
            self._assign_thread.deleteLater()
            self._assign_thread = None
        self.sidebar.setEnabled(True)
        if self.filter_tab:
            self.filter_tab._apply_btn.setEnabled(True)

    def _on_assignment_finished(self, has_items: bool):
        ctrl = self.controller
        write_debug("qt.assignment_finished",
                     has_items=has_items,
                     filtered_count=len(ctrl.session.filtered_items) if has_items else 0)

        if not has_items:
            QMessageBox.warning(
                self, "No Items",
                "After applying filters, no items remain to order.",
            )
            return

        # Clear undo stack for fresh session
        self.undo_stack.clear()

        # Populate the bulk grid
        if self.bulk_tab is not None:
            self.bulk_tab.set_data(
                ctrl.session.filtered_items,
                ctrl.session.inventory_lookup,
                ctrl.order_rules,
                ctrl._suggest_min_max,
            )
            self.bulk_tab.set_known_vendors(ctrl.vendor_codes_used)

            self._mark_step("Filter", "done")
            self._mark_step("Bulk", "active")

            # Switch to Bulk tab
            bulk_idx = next(
                (i for i, (label, _) in enumerate(NAV_ITEMS) if label == "Bulk"),
                None,
            )
            if bulk_idx is not None:
                self.sidebar.setCurrentRow(bulk_idx)

        auto_result = getattr(ctrl, "_last_auto_assign_result", {})
        assigned_count = auto_result.get("assigned_count", 0)

        items = ctrl.session.filtered_items
        item_count = len(items)
        assigned_total = sum(1 for i in items if i.get("vendor"))
        unassigned = item_count - assigned_total
        review_count = sum(1 for i in items if str(i.get("status", "")).lower() == "review")
        warning_count = sum(1 for i in items if str(i.get("status", "")).lower() in ("warning", "warn"))
        dead_count = sum(1 for i in items if i.get("dead_stock"))

        self._status_item_count.setText(
            f"{item_count:,} items \u00b7 {assigned_total:,} assigned"
        )

        # Build summary lines
        lines = [f"<b>{item_count:,}</b> items loaded"]
        if assigned_count:
            lines.append(f"<b>{assigned_count:,}</b> auto-assigned from receipt history")
        if unassigned:
            lines.append(f"<b>{unassigned:,}</b> need manual vendor assignment")
        if review_count:
            lines.append(f"<b>{review_count:,}</b> need review (reel/manual/pack)")
        if warning_count:
            lines.append(f"<b>{warning_count:,}</b> warnings (missing data / deferred)")
        if dead_count:
            lines.append(f"<b>{dead_count:,}</b> dead stock items")

        body = "<br>".join(lines)
        QMessageBox.information(
            self, "Session Ready",
            f"<p>{body}</p>"
            f"<p style='color: gray;'>Use the bulk grid to assign remaining vendors, "
            f"then Review & Export when ready.</p>",
        )
        self.status.showMessage(f"{item_count:,} items ready", 15000)

    def _on_assignment_failed(self, error: str):
        write_debug("qt.assignment_failed", error=error)
        QMessageBox.critical(
            self, "Assignment Error",
            f"Could not prepare vendor assignment:\n{error}",
        )

    # ── Bulk grid editing ─────────────────────────────────────────

    def _on_cell_edit(self, item: dict, col_name: str, value: str):
        """Handle a cell edit from the bulk grid delegate."""
        ctrl = self.controller
        lc = item.get("line_code", "")
        ic = item.get("item_code", "")
        key = (lc, ic)
        inv = ctrl.session.inventory_lookup.get(key, {})

        write_debug("qt.cell_edit.begin",
                     col_name=col_name, value=str(value)[:40],
                     line_code=lc, item_code=ic)

        if col_name == "vendor":
            old_val = item.get("vendor", "")
            new_val = value.strip().upper()
            if new_val:
                item["vendor"] = new_val
                ctrl._remember_vendor_code(new_val)
                if self.bulk_tab:
                    self.bulk_tab.set_known_vendors(ctrl.vendor_codes_used)
            write_debug("qt.cell_edit.vendor",
                         line_code=lc, item_code=ic,
                         old_vendor=old_val, new_vendor=new_val)

        elif col_name == "notes":
            old_notes = item.get("notes", "")
            item["notes"] = value.strip()
            write_debug("qt.cell_edit.notes",
                         line_code=lc, item_code=ic,
                         old_notes=old_notes[:30], new_notes=value.strip()[:30])
            try:
                import item_notes_flow
                nk = f"{lc}:{ic}"
                if value.strip():
                    ctrl.item_notes[nk] = value.strip()
                else:
                    ctrl.item_notes.pop(nk, None)
                item_notes_flow.save_notes(ctrl._data_path("item_notes"), ctrl.item_notes)
                write_debug("qt.cell_edit.notes.saved", key=nk)
            except Exception as exc:
                write_debug("qt.cell_edit.notes.save_error", error=str(exc))

        elif col_name == "final_qty":
            try:
                qty = int(float(value))
                old_final = item.get("final_qty", 0)
                old_suggested = item.get("suggested_qty", 0)
                ctrl._set_effective_order_qty(item, qty, manual_override=True)
                ctrl._recalculate_item(item)
                item.pop("_text_haystack", None)
                write_debug("qt.cell_edit.final_qty",
                             line_code=lc, item_code=ic,
                             old_final=old_final, new_final=item.get("final_qty"),
                             old_suggested=old_suggested, new_suggested=item.get("suggested_qty"),
                             manual_override=item.get("manual_override"))
            except (ValueError, TypeError) as exc:
                write_debug("qt.cell_edit.final_qty.error", line_code=lc, item_code=ic, error=str(exc))
                return

        elif col_name == "qoh":
            try:
                new_qoh = float(value)
                old_qoh = inv.get("qoh", 0)
                if new_qoh != old_qoh:
                    ctrl.qoh_adjustments[key] = {"old": old_qoh, "new": new_qoh}
                    if key in ctrl.session.inventory_lookup:
                        ctrl.session.inventory_lookup[key]["qoh"] = new_qoh
                    ctrl._recalculate_item(item)
                    item.pop("_text_haystack", None)
                    write_debug("qt.cell_edit.qoh",
                                 line_code=lc, item_code=ic,
                                 old_qoh=old_qoh, new_qoh=new_qoh,
                                 new_suggested=item.get("suggested_qty"),
                                 new_final=item.get("final_qty"))
            except (ValueError, TypeError) as exc:
                write_debug("qt.cell_edit.qoh.error", line_code=lc, item_code=ic, error=str(exc))
                return

        elif col_name in ("cur_min", "cur_max"):
            try:
                parsed = None if value == "" else int(float(value))
                if key not in ctrl.session.inventory_lookup:
                    ctrl.session.inventory_lookup[key] = {
                        "qoh": 0, "repl_cost": 0, "min": None, "max": None,
                        "ytd_sales": 0, "mo12_sales": 0, "supplier": "",
                        "last_receipt": "", "last_sale": "",
                    }
                old_val = ctrl.session.inventory_lookup[key].get("min" if col_name == "cur_min" else "max")
                if col_name == "cur_min":
                    ctrl.session.inventory_lookup[key]["min"] = parsed
                else:
                    ctrl.session.inventory_lookup[key]["max"] = parsed
                ctrl._recalculate_item(item)
                item.pop("_text_haystack", None)
                write_debug("qt.cell_edit.minmax",
                             line_code=lc, item_code=ic, col=col_name,
                             old_value=old_val, new_value=parsed,
                             new_suggested=item.get("suggested_qty"),
                             new_final=item.get("final_qty"))
            except (ValueError, TypeError) as exc:
                write_debug("qt.cell_edit.minmax.error", line_code=lc, item_code=ic, error=str(exc))
                return

        elif col_name == "pack_size":
            try:
                old_pack = item.get("pack_size")
                new_pack = int(float(value)) if value else None
                item["pack_size"] = new_pack
                if new_pack:
                    item["pack_size_source"] = "manual"
                rule_key = f"{lc}:{ic}"
                rule = ctrl.order_rules.get(rule_key) or {}
                if new_pack:
                    rule["pack_size"] = new_pack
                    ctrl.order_rules[rule_key] = rule
                else:
                    rule.pop("pack_size", None)
                    if not rule:
                        ctrl.order_rules.pop(rule_key, None)
                    else:
                        ctrl.order_rules[rule_key] = rule
                ctrl._save_order_rules()
                old_suggested = item.get("suggested_qty")
                old_final = item.get("final_qty")
                ctrl._recalculate_item(item)
                item.pop("_text_haystack", None)
                write_debug("qt.cell_edit.pack_size",
                             line_code=lc, item_code=ic,
                             old_pack=old_pack, new_pack=new_pack,
                             old_suggested=old_suggested, new_suggested=item.get("suggested_qty"),
                             old_final=old_final, new_final=item.get("final_qty"),
                             rule_saved=True)
            except Exception as exc:
                write_debug("qt.cell_edit.pack_size.error", line_code=lc, item_code=ic, error=str(exc))
                return

    def _on_cycle_changed(self, weeks: int):
        """Re-enrich all items when the reorder cycle changes."""
        write_debug("qt.cycle_changed", weeks=weeks)
        ctrl = self.controller
        ctrl._cycle_weeks = weeks
        ctrl._suggest_min_max_cache = {}

        if not self.bulk_tab or not ctrl.session.filtered_items:
            return

        import time
        t0 = time.perf_counter()
        # Re-normalize and re-enrich all items
        reorder_flow = __import__("reorder_flow")
        reorder_flow.normalize_items_to_cycle(ctrl)
        elapsed = (time.perf_counter() - t0) * 1000
        write_debug("qt.cycle_changed.renormalize", elapsed_ms=round(elapsed, 1),
                     items=len(ctrl.session.filtered_items))

        self.bulk_tab.model.bump_generation()
        self.bulk_tab.model.set_data(
            ctrl.session.filtered_items,
            ctrl.session.inventory_lookup,
            ctrl.order_rules,
            ctrl._suggest_min_max,
        )
        self.bulk_tab._update_summary()
        self.status.showMessage(f"Recalculated for {weeks}-week cycle", 5000)

    def _on_draft_review(self):
        """Export draft review print files — per-vendor landscape xlsx for physical verification."""
        write_debug("qt.draft_review.begin")
        ctrl = self.controller
        model = self.bulk_tab.model if self.bulk_tab else None
        if not model or not model.items:
            QMessageBox.information(self, "No Data", "Load data and assign vendors first.")
            return

        output_dir = QFileDialog.getExistingDirectory(
            self, "Select Folder for Draft Review Files",
        )
        if not output_dir:
            return

        try:
            import draft_report_flow
            created = draft_report_flow.export_draft_review_files(
                model.items,
                ctrl.session.inventory_lookup,
                output_dir,
                receipt_cost_lookup=getattr(ctrl.session, "receipt_cost_lookup", None),
            )
            if created:
                file_list = "\n".join(f"  \u2022 {vendor}: {os.path.basename(path)}"
                                      for vendor, path in created[:10])
                if len(created) > 10:
                    file_list += f"\n  \u2026 and {len(created) - 10} more"
                QMessageBox.information(
                    self, "Draft Review Complete",
                    f"Created {len(created)} draft review file(s) in:\n{output_dir}\n\n{file_list}",
                )
                write_debug("qt.draft_review.done", files=len(created))
            else:
                QMessageBox.information(
                    self, "No Vendors",
                    "No assigned items with order quantities to review.",
                )
        except Exception as exc:
            write_debug("qt.draft_review.error", error=str(exc))
            QMessageBox.critical(
                self, "Draft Review Error", f"Failed to generate draft review:\n{exc}",
            )

    # ── Settings ──────────────────────────────────────────────────

    def _on_settings(self):
        from ui_qt.settings_dialog import SettingsDialog
        dialog = SettingsDialog(self.app_settings, parent=self)
        if dialog.exec() == SettingsDialog.Accepted and dialog.changed:
            self._save_settings()
            write_debug("qt.settings.saved")

    def _save_settings(self):
        import storage
        import sys as _sys
        if getattr(_sys, "frozen", False):
            data_dir = os.path.dirname(_sys.executable)
        else:
            data_dir = os.path.dirname(os.path.abspath(__file__))
            data_dir = os.path.dirname(data_dir)  # up from ui_qt/
        path = os.path.join(data_dir, "po_builder_settings.json")
        try:
            storage.save_json_file(path, self.app_settings)
        except Exception as exc:
            write_debug("qt.settings.save_error", error=str(exc))

    # ── Analysis Reports ───────────────────────────────────────────

    def _on_export_dead_stock(self):
        model = self.bulk_tab.model if self.bulk_tab else None
        if not model or not model.items:
            QMessageBox.information(self, "No Data", "Load data first.")
            return
        output_dir = QFileDialog.getExistingDirectory(self, "Select Folder for Dead Stock Report")
        if not output_dir:
            return
        import analysis_reports
        inv = self.controller.session.inventory_lookup or {}
        path = analysis_reports.export_dead_stock_xlsx(model.items, inv, output_dir)
        summary = analysis_reports.dead_stock_summary(
            analysis_reports.build_dead_stock_rows(model.items, inv)
        )
        QMessageBox.information(
            self, "Dead Stock Report",
            f"Exported {summary['total_items']} dead stock items "
            f"(${summary['total_on_hand_value']:,.0f} on-hand value) "
            f"across {len(summary['vendors'])} vendor(s).\n\n{os.path.basename(path)}",
        )

    def _on_export_deferred(self):
        model = self.bulk_tab.model if self.bulk_tab else None
        if not model or not model.items:
            QMessageBox.information(self, "No Data", "Load data first.")
            return
        output_dir = QFileDialog.getExistingDirectory(self, "Select Folder for Deferred Items Report")
        if not output_dir:
            return
        import analysis_reports
        inv = self.controller.session.inventory_lookup or {}
        rows = analysis_reports.build_deferred_rows(model.items, inv)
        if not rows:
            QMessageBox.information(self, "No Deferred Items", "No items were deferred this session.")
            return
        path = analysis_reports.export_deferred_csv(model.items, inv, output_dir)
        QMessageBox.information(
            self, "Deferred Items Report",
            f"Exported {len(rows)} deferred item(s).\n\n{os.path.basename(path)}",
        )

    def _on_export_session_summary(self):
        model = self.bulk_tab.model if self.bulk_tab else None
        if not model or not model.items:
            QMessageBox.information(self, "No Data", "Load data first.")
            return
        output_dir = QFileDialog.getExistingDirectory(self, "Select Folder for Session Summary")
        if not output_dir:
            return
        import analysis_reports
        inv = self.controller.session.inventory_lookup or {}
        path = analysis_reports.export_session_summary_csv(model.items, inv, output_dir)
        summary = analysis_reports.build_session_summary(model.items, inv)
        QMessageBox.information(
            self, "Session Summary",
            f"Exported session summary:\n"
            f"  {summary['total_items']:,} items, {summary['assigned']:,} assigned\n"
            f"  Est. order value: ${summary['total_order_value']:,.0f}\n"
            f"  {len(summary['vendor_summaries'])} vendor(s)\n\n"
            f"{os.path.basename(path)}",
        )

    def _on_before_cell_edit(self, row: int, col_name: str):
        """Capture before-state for undo on cell edits."""
        write_debug("qt.before_cell_edit", row=row, col_name=col_name)
        if self.bulk_tab:
            self.undo_stack.push_edit(
                f"edit:{col_name}",
                self.bulk_tab.model.items,
                [row],
            )

    def _on_undo(self):
        if not self.bulk_tab:
            return
        model = self.bulk_tab.model
        entry = self.undo_stack.undo(model.items)
        if entry is None:
            self.status.showMessage("Nothing to undo", 3000)
            return
        write_debug("qt.undo", label=entry.label,
                     snapshots=len(entry.item_snapshots),
                     removals=len(entry.removed_items))
        # Refresh the model
        model.bump_generation()
        if entry.removed_items:
            # Re-inserted items — full reset
            model.set_data(
                model.items,
                self.controller.session.inventory_lookup,
                self.controller.order_rules,
                self.controller._suggest_min_max,
            )
        elif entry.item_snapshots:
            model.refresh_rows(list(entry.item_snapshots.keys()))
        self.bulk_tab._update_summary()
        self.status.showMessage(f"Undone: {entry.label}", 3000)

    def _on_redo(self):
        if not self.bulk_tab:
            return
        model = self.bulk_tab.model
        entry = self.undo_stack.redo(model.items)
        if entry is None:
            self.status.showMessage("Nothing to redo", 3000)
            return
        write_debug("qt.redo", label=entry.label,
                     snapshots=len(entry.item_snapshots),
                     removals=len(entry.removed_items))
        model.bump_generation()
        if entry.removed_items:
            model.set_data(
                model.items,
                self.controller.session.inventory_lookup,
                self.controller.order_rules,
                self.controller._suggest_min_max,
            )
        elif entry.item_snapshots:
            model.refresh_rows(list(entry.item_snapshots.keys()))
        self.bulk_tab._update_summary()
        self.status.showMessage(f"Redone: {entry.label}", 3000)

    def _on_vendor_applied(self, source_rows: list, vendor: str):
        """Apply a vendor code to a set of rows."""
        write_debug("qt.vendor_apply.begin", vendor=vendor, row_count=len(source_rows))
        ctrl = self.controller
        model = self.bulk_tab.model if self.bulk_tab else None
        # Capture before-state for undo
        if model:
            self.undo_stack.push_edit(f"vendor:{vendor}", model.items, source_rows)
        changed_rows = []
        for row in source_rows:
            item = model.item_at(row) if model else None
            if item is not None:
                item["vendor"] = vendor
                item.pop("_text_haystack", None)
                changed_rows.append(row)
        ctrl._remember_vendor_code(vendor)

        if self.bulk_tab:
            self.bulk_tab.model.bump_generation()
            self.bulk_tab.model.refresh_rows(changed_rows)
            self.bulk_tab.set_known_vendors(ctrl.vendor_codes_used)
            self.bulk_tab._update_summary()

        self.status.showMessage(
            f"Applied vendor {vendor} to {len(changed_rows)} item(s)", 5000,
        )
        write_debug("qt.vendor_apply.done", vendor=vendor, changed=len(changed_rows))

    def _on_rows_removed(self, source_rows: list):
        """Remove items from the session."""
        write_debug("qt.rows_removed.begin", row_count=len(source_rows))
        ctrl = self.controller
        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return
        items = model.items
        # Capture for undo before removing
        removal_pairs = []
        for row in sorted(set(source_rows), reverse=True):
            if 0 <= row < len(items):
                removal_pairs.append((row, items[row]))
        self.undo_stack.push_removal("remove:rows", removal_pairs)
        # Remove in reverse order to preserve indices
        removed = []
        for row in sorted(set(source_rows), reverse=True):
            if 0 <= row < len(items):
                removed.append(items.pop(row))
        ctrl.last_removed_bulk_items = removed

        if self.bulk_tab:
            self.bulk_tab.model.set_data(
                items,
                ctrl.session.inventory_lookup,
                ctrl.order_rules,
                ctrl._suggest_min_max,
            )
            self.bulk_tab._update_summary()

        write_debug("qt.rows_removed.done",
                     removed=len(removed),
                     remaining=model.rowCount() if model else 0)
        self.status.showMessage(
            f"Removed {len(removed)} item(s)", 5000,
        )

    # ── Export ────────────────────────────────────────────────────

    def _on_export_requested(self, _output_dir: str):
        """Write per-vendor xlsx files through the shared export flow.

        Runs stock warnings check first — mirrors tkinter's
        check_stock_warnings gate before export.
        """
        if not self._qt_check_stock_warnings():
            return
        # Capture PO memo from review tab before exporting
        self._po_memo = ""
        if self.review_tab and hasattr(self.review_tab, "_memo_edit"):
            self._po_memo = self.review_tab._memo_edit.text().strip()
        assigned = self._assigned_items_for_export()
        write_debug("qt.export.begin", assigned=len(assigned),
                     po_memo=bool(self._po_memo))
        export_flow.do_export(
            self,
            self._export_vendor_po,
            self.controller._data_path("order_history"),
            self.controller._data_path("sessions"),
            assigned_items=assigned,
        )
        self._mark_step("Bulk", "done")
        self._mark_step("Review", "done")
        return

    def _qt_check_stock_warnings(self) -> bool:
        """Pre-export stock warnings check. Returns True to proceed."""
        import perf_trace
        from rules.not_needed import not_needed_reason
        ctrl = self.controller
        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return True

        self._show_loading("Checking stock warnings\u2026")
        QApplication.processEvents()

        flagged = []
        with perf_trace.span("qt.check_stock_warnings.scan",
                              item_count=len(model.items)):
            for item in model.items:
                if not item.get("vendor"):
                    continue
                key = (item.get("line_code", ""), item.get("item_code", ""))
                inv = ctrl.session.inventory_lookup.get(key, {})
                try:
                    reason_text, _ = not_needed_reason(ctrl, item, max_exceed_abs_buffer=5)
                except Exception:
                    reason_text = ""
                if reason_text:
                    reasons = [p.strip() for p in reason_text.split(";") if p.strip()]
                    _, sug_max = ctrl._suggest_min_max(key)
                    flagged.append({
                        "item": item,
                        "qoh": inv.get("qoh", 0),
                        "min": inv.get("min"),
                        "max": inv.get("max"),
                        "sug_max": sug_max,
                        "pack_size": item.get("pack_size"),
                        "reasons": reasons,
                    })

        self._hide_loading()
        write_debug("qt.check_stock_warnings.done", flagged=len(flagged))

        if not flagged:
            return True

        MAX_FLAGGED_ROWS = 50
        if len(flagged) > MAX_FLAGGED_ROWS:
            from ui_qt.workflow_dialogs import TooManyFlaggedDialog
            dialog = TooManyFlaggedDialog(len(flagged), parent=self)
            dialog.exec()
            return dialog.proceed

        from ui_qt.workflow_dialogs import StockWarningsDialog
        dialog = StockWarningsDialog(flagged, parent=self)
        dialog.exec()
        if not dialog.proceed:
            return False
        # Clear vendor on unchecked items
        for item in dialog.items_to_unassign:
            real_item = item.get("item", item)
            real_item["vendor"] = ""
        if dialog.items_to_unassign:
            model.set_data(model.items, ctrl.session.inventory_lookup,
                           ctrl.order_rules, ctrl._suggest_min_max)
            self.bulk_tab._update_summary()
        return True

    def _assigned_items_for_export(self) -> list[dict]:
        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return []
        return [item for item in model.items if str(item.get("vendor", "")).strip()]

    def _loaded_report_paths_for_snapshot(self) -> dict:
        return dict(self._loaded_report_paths)

    def _show_warning(self, title: str, message: str):
        return QMessageBox.warning(self, title, message)

    def _show_info(self, title: str, message: str):
        return QMessageBox.information(self, title, message)

    def _show_error(self, title: str, message: str):
        return QMessageBox.critical(self, title, message)

    def _ask_yes_no(self, title: str, message: str) -> bool:
        result = QMessageBox.question(self, title, message, QMessageBox.Yes | QMessageBox.No)
        return result == QMessageBox.Yes

    def _ask_yes_no_cancel(self, title: str, message: str):
        result = QMessageBox.question(
            self,
            title,
            message,
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
        )
        if result == QMessageBox.Cancel:
            return None
        return result == QMessageBox.Yes

    def _get_mixed_export_behavior(self) -> str:
        return str(self.app_settings.get("mixed_export_behavior", "all_exportable") or "").strip() or "all_exportable"

    def _get_planned_only_export_behavior(self) -> str:
        return (
            str(self.app_settings.get("planned_only_export_behavior", "export_automatically") or "").strip()
            or "export_automatically"
        )

    def _get_last_export_dir(self) -> str:
        return str(self.app_settings.get("last_export_dir", "") or "").strip()

    def _set_last_export_dir(self, path: str):
        normalized = str(path or "").strip()
        if normalized:
            self.app_settings["last_export_dir"] = normalized

    def _choose_output_dir(self) -> str:
        start = self._get_last_export_dir()
        output_dir = QFileDialog.getExistingDirectory(
            self,
            "Select Output Folder for PO Files",
            start,
        )
        if output_dir:
            self._set_last_export_dir(output_dir)
        return output_dir

    def _show_loading(self, text="Working..."):
        self._loading_depth += 1
        self._loading_message = text
        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.status.showMessage(text)

    def _hide_loading(self):
        if self._loading_depth > 0:
            self._loading_depth -= 1
        while QApplication.overrideCursor() is not None:
            QApplication.restoreOverrideCursor()
        if self._loading_depth == 0:
            self.status.clearMessage()

    def _process_events(self):
        QApplication.processEvents()

    def _show_export_preview_dialog(self, preview_data: dict) -> bool:
        from ui_qt.export_dialogs import ExportPreviewDialog

        dialog = ExportPreviewDialog(
            preview_data,
            initial_overrides=self._vendor_export_scope_overrides,
            parent=self,
        )
        accepted = dialog.exec() == ExportPreviewDialog.Accepted
        if accepted:
            self._vendor_export_scope_overrides = dialog.overrides
        return accepted

    def _get_x4_pack_size(self, key):
        pack_lookup = getattr(self.controller.session, "pack_size_source_lookup", {}) or {}
        pack = pack_lookup.get(key)
        if pack:
            return pack
        generic = pack_lookup.get(("", key[1]))
        if generic:
            return generic
        return None

    def _build_maintenance_report(self):
        candidates = maintenance_flow.build_maintenance_candidates(
            self.controller.session,
            suggest_min_max=self.controller._suggest_min_max,
            get_x4_pack_size=self._get_x4_pack_size,
        )
        return maintenance.build_maintenance_report(candidates)

    def _show_maintenance_report(self, output_dir, issues=None):
        issues = list(issues or [])
        if not issues:
            self.status.showMessage(f"Exported to {output_dir}", 10000)
            return
        csv_path = export_flow.export_maintenance_csv(issues, output_dir)
        QMessageBox.information(
            self,
            "Maintenance Report",
            f"Generated {len(issues)} maintenance item(s).\n\nCSV: {csv_path}",
        )

    def _export_vendor_po(self, vendor: str, items: list[dict], output_dir: str):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from datetime import datetime

        po_memo = getattr(self, "_po_memo", "") or ""
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PO Import"

        headers = ["product group", "item code", "order quantity"]
        has_item_notes = any(item.get("notes") for item in items)
        include_notes = bool(po_memo) or has_item_notes
        if include_notes:
            headers.append("Notes")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            pg = ws.cell(row=row_idx, column=1, value=item.get("line_code", ""))
            ic = ws.cell(row=row_idx, column=2, value=item.get("item_code", ""))
            qty = ws.cell(row=row_idx, column=3, value=item.get("final_qty", item.get("order_qty", 0)))
            for cell in (pg, ic, qty):
                cell.border = thin_border
            ic.number_format = "@"
            qty.alignment = Alignment(horizontal="center")
            if include_notes:
                note_text = item.get("notes", "") or po_memo
                note = ws.cell(row=row_idx, column=4, value=note_text)
                note.border = thin_border

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 16
        if has_notes:
            ws.column_dimensions["D"].width = 32

        safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in vendor)
        timestamp = datetime.now().strftime("%Y%m%d")
        filepath = os.path.join(output_dir, f"PO_{safe_name}_{timestamp}.xlsx")
        wb.save(filepath)
        write_debug("qt.export.vendor_file", vendor=vendor, items=len(items), file=os.path.basename(filepath))
        return filepath

    def _on_remove_not_needed(self):
        """Remove items that don't need ordering — scan + interactive dialog.

        The scan takes ~730ms on 56K items — fast enough to run synchronously
        with processEvents for status updates.  QThread was tried but caused
        signal delivery issues (callback never fired on the main thread).
        """
        import time as _time
        import perf_trace
        write_debug("qt.remove_not_needed.begin")
        ctrl = self.controller
        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return

        self._show_loading("Scanning for not-needed items\u2026")
        QApplication.processEvents()

        from rules.not_needed import not_needed_reason

        items = model.items
        candidates = []
        skipped_assigned = 0
        total = len(items)
        t0 = _time.perf_counter()
        perf_trace.stamp("qt.remove_not_needed.scan.begin", total=total)
        report_interval = max(1, total // 20)
        for i, item in enumerate(items):
            if i % report_interval == 0 and i > 0:
                self.status.showMessage(
                    f"Scanning\u2026 {i:,}/{total:,} ({len(candidates)} flagged)"
                )
                QApplication.processEvents()
            vendor = str(item.get("vendor", "") or "").strip()
            if vendor:
                item_needs_order = (
                    (item.get("final_qty") or 0) > 0
                    and item.get("status") != "skip"
                )
                if item_needs_order:
                    skipped_assigned += 1
                    continue
            try:
                result = not_needed_reason(ctrl, item, max_exceed_abs_buffer=50)
                reason = result[0] if isinstance(result, (list, tuple)) else result
                auto = (result[1]
                        if isinstance(result, (list, tuple)) and len(result) > 1
                        else bool(reason))
                if reason:
                    key = (item.get("line_code", ""), item.get("item_code", ""))
                    inv = ctrl.session.inventory_lookup.get(key, {})
                    _, sug_max = ctrl._suggest_min_max(key)
                    candidates.append({
                        "index": i,
                        "line_code": item.get("line_code", ""),
                        "item_code": item.get("item_code", ""),
                        "description": item.get("description", ""),
                        "final_qty": item.get("final_qty", item.get("order_qty", 0)),
                        "qoh": inv.get("qoh", 0),
                        "max": inv.get("max"),
                        "sug_max": sug_max,
                        "reason": reason,
                        "auto_remove": auto,
                    })
            except Exception as exc:
                write_debug("qt.remove_not_needed.item_error",
                             index=i,
                             line_code=item.get("line_code", ""),
                             item_code=item.get("item_code", ""),
                             error=str(exc))
        elapsed_ms = (_time.perf_counter() - t0) * 1000
        perf_trace.stamp("qt.remove_not_needed.scan.done",
                          total=total, flagged=len(candidates),
                          skipped_assigned=skipped_assigned,
                          elapsed_ms=round(elapsed_ms, 1))
        self._hide_loading()
        write_debug("qt.remove_not_needed.scanned",
                     total=total, flagged=len(candidates),
                     skipped_assigned=skipped_assigned)
        self._show_remove_not_needed_dialog(
            items, candidates, skipped_assigned, model, ctrl,
        )

    def _show_remove_not_needed_dialog(
        self, items, candidates, skipped_assigned, model, ctrl,
    ):
        """Open the interactive dialog after scanning completes."""
        if not candidates:
            QMessageBox.information(self, "Nothing to Remove",
                                    "All items appear to need ordering.")
            return

        try:
            write_debug("qt.remove_not_needed.dialog.build",
                         candidates=len(candidates))
            from ui_qt.workflow_dialogs import RemoveNotNeededDialog
            dialog = RemoveNotNeededDialog(
                candidates,
                excluded_assigned_count=skipped_assigned,
                parent=self,
            )
            write_debug("qt.remove_not_needed.dialog.show")
            if dialog.exec() != RemoveNotNeededDialog.Accepted:
                write_debug("qt.remove_not_needed.dialog.cancelled")
                return
        except Exception as exc:
            write_debug("qt.remove_not_needed.dialog.error", error=str(exc))
            QMessageBox.critical(self, "Error", f"Dialog failed: {exc}")
            return

        removed_candidate_indices = dialog.removed_indices
        if not removed_candidate_indices:
            return

        removal_indices = sorted(
            [candidates[ci]["index"] for ci in removed_candidate_indices],
            reverse=True,
        )

        # Capture for undo
        removal_pairs = [(i, items[i]) for i in removal_indices]
        self.undo_stack.push_removal("remove:not_needed", removal_pairs)

        for i in removal_indices:
            items.pop(i)

        model.set_data(
            items,
            ctrl.session.inventory_lookup,
            ctrl.order_rules,
            ctrl._suggest_min_max,
        )
        self.bulk_tab._update_summary()
        self.status.showMessage(
            f"Removed {len(removal_indices)} not-needed item(s)", 5000,
        )
        write_debug("qt.remove_not_needed.done",
                     removed=len(removal_indices),
                     remaining=model.rowCount() if model else 0)

    # ── Buy Rule Edit Committed ──────────────────────────────────

    def _on_edit_committed(self, source_row: int, col_name: str, value: str):
        """Handle buy rule or other committed edits from the bulk tab."""
        write_debug("qt.edit_committed", col_name=col_name, value=value, source_row=source_row)
        ctrl = self.controller
        if col_name == "buy_rule":
            ctrl._save_order_rules()
            rule = ctrl.order_rules.get(value, {})
            write_debug("qt.buy_rule.saved", rule_key=value,
                         pack_size=rule.get("pack_size"),
                         trigger_qty=rule.get("reorder_trigger_qty"),
                         min_packs=rule.get("minimum_packs_on_hand"),
                         cover_days=rule.get("cover_days"),
                         exact_qty=rule.get("exact_order_qty"))
            if source_row >= 0:
                model = self.bulk_tab.model if self.bulk_tab else None
                if model:
                    item = model.item_at(source_row)
                    if item:
                        # Apply the rule's pack_size to the item before recalculating
                        # (mirrors assignment_flow lines 412-414)
                        from rules import get_rule_pack_size, has_exact_qty_override
                        rule_pack = get_rule_pack_size(rule)
                        if rule_pack is not None:
                            item["pack_size"] = rule_pack
                            item["pack_size_source"] = "rule"
                        elif has_exact_qty_override(rule):
                            item["pack_size"] = None
                            item["pack_size_source"] = "rule_exact_qty"
                        old_sug = item.get("suggested_qty")
                        old_final = item.get("final_qty")
                        ctrl._clear_manual_override(item)
                        ctrl._recalculate_item(item)
                        item.pop("_text_haystack", None)
                        model.bump_generation()
                        model.refresh_rows([source_row])
                        write_debug("qt.buy_rule.recalculated",
                                     rule_key=value,
                                     old_suggested=old_sug, new_suggested=item.get("suggested_qty"),
                                     old_final=old_final, new_final=item.get("final_qty"),
                                     new_pack=item.get("pack_size"),
                                     new_policy=item.get("order_policy"))
            self.status.showMessage("Buy rule saved", 3000)

    # ── Ignore Item ───────────────────────────────────────────────

    def _on_ignore_item(self, line_code: str, item_code: str):
        """Add item to the ignore list and remove from grid."""
        ctrl = self.controller
        ignore_key = f"{line_code}:{item_code}"
        write_debug("qt.ignore_item", key=ignore_key)
        ctrl.ignored_item_keys.add(ignore_key)
        try:
            import storage
            storage.save_ignored_items(
                ctrl._data_path("ignored_items"), ctrl.ignored_item_keys,
            )
        except Exception:
            pass

        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return

        # Find and remove the item
        items = model.items
        removal_idx = None
        for i, item in enumerate(items):
            if item.get("line_code") == line_code and item.get("item_code") == item_code:
                removal_idx = i
                break

        if removal_idx is not None:
            self.undo_stack.push_removal(
                f"ignore:{ignore_key}",
                [(removal_idx, items[removal_idx])],
            )
            items.pop(removal_idx)
            model.set_data(
                items,
                ctrl.session.inventory_lookup,
                ctrl.order_rules,
                ctrl._suggest_min_max,
            )
            self.bulk_tab._update_summary()

        self.status.showMessage(f"Ignored {ignore_key}", 5000)

    # ── Workflow dialog launchers ──────────────────────────────────

    def _on_vendor_review(self):
        """Open vendor review dialog with session snapshot data."""
        import storage
        import vendor_summary_flow
        ctrl = self.controller
        sessions_dir = ctrl._data_path("sessions")
        snapshots = storage.load_session_snapshots(sessions_dir, max_count=25) or []
        lead_times = storage.infer_vendor_lead_times(snapshots) or {}
        vendor_codes = list(getattr(ctrl, "vendor_codes_used", None) or [])
        summaries = vendor_summary_flow.summarize_all_vendors(
            snapshots, vendor_codes=vendor_codes or None,
            lead_times=lead_times, top_n=5,
        )
        from ui_qt.workflow_dialogs import VendorReviewDialog
        dialog = VendorReviewDialog(summaries, parent=self)
        dialog.exec()

    def _on_session_diff(self):
        """Open session diff dialog."""
        import session_diff_flow
        ctrl = self.controller
        sessions_dir = ctrl._data_path("sessions")
        previous = session_diff_flow.load_previous_snapshot(sessions_dir)
        items = []
        model = self.bulk_tab.model if self.bulk_tab else None
        if model:
            items = list(model.items)
        current = {"exported_items": items}
        diff = session_diff_flow.diff_sessions(previous, current)
        summary_text = session_diff_flow.format_diff_summary(diff) or "No changes since the last session."
        snap_label = session_diff_flow.snapshot_label(previous) or ""
        from ui_qt.workflow_dialogs import SessionDiffDialog
        dialog = SessionDiffDialog(diff, summary_text=summary_text,
                                    snapshot_label=snap_label, parent=self)
        dialog.exec()

    def _on_supplier_map(self):
        """Open supplier map editor."""
        import supplier_map_flow
        ctrl = self.controller
        path = ctrl._data_path("supplier_vendor_map")
        mapping = supplier_map_flow.load_supplier_map(path)

        from ui_qt.workflow_dialogs import SupplierMapDialog
        dialog = SupplierMapDialog(mapping, parent=self)

        def _on_learn():
            import storage
            sessions_dir = ctrl._data_path("sessions")
            snapshots = storage.load_session_snapshots(sessions_dir, max_count=25) or []
            inferred = supplier_map_flow.build_supplier_map_from_history(snapshots)
            if not inferred:
                QMessageBox.information(dialog, "Nothing to Learn",
                                        "No supplier\u2192vendor pairs found in snapshots.")
                return
            merged = supplier_map_flow.merge_supplier_maps(dialog.working_mapping, inferred)
            dialog.set_mapping(merged)

        def _on_apply(working):
            model = self.bulk_tab.model if self.bulk_tab else None
            if model is None:
                return
            pairs = supplier_map_flow.apply_supplier_map(model.items, working)
            if not pairs:
                QMessageBox.information(dialog, "Nothing to Apply",
                                        "No unassigned items match a mapped supplier.")
                return
            for item, vendor in pairs:
                item["vendor"] = vendor
            model.set_data(model.items, ctrl.session.inventory_lookup,
                           ctrl.order_rules, ctrl._suggest_min_max)
            self.bulk_tab._update_summary()
            QMessageBox.information(dialog, "Applied",
                                    f"Auto-assigned {len(pairs)} item(s).")

        def _on_save(working):
            try:
                supplier_map_flow.save_supplier_map(path, working)
            except Exception as exc:
                QMessageBox.critical(dialog, "Save Failed", str(exc))

        dialog.learn_requested.connect(_on_learn)
        dialog.apply_requested.connect(_on_apply)
        dialog.save_requested.connect(_on_save)
        dialog.exec()

    def _on_qoh_review(self):
        """Open QOH adjustments review dialog."""
        import qoh_review_flow
        ctrl = self.controller
        adjustments = getattr(ctrl, "qoh_adjustments", {}) or {}
        inv_lookup = ctrl.session.inventory_lookup if ctrl.session else {}
        rows = qoh_review_flow.format_qoh_adjustments(adjustments, inv_lookup)

        from ui_qt.workflow_dialogs import QohReviewDialog
        dialog = QohReviewDialog(rows, parent=self)
        if dialog.exec() == QohReviewDialog.Accepted and dialog.reverted_keys:
            reverted = qoh_review_flow.revert_qoh_adjustments(
                adjustments, inv_lookup, dialog.reverted_keys,
            )
            if reverted:
                model = self.bulk_tab.model if self.bulk_tab else None
                if model:
                    model.set_data(model.items, inv_lookup,
                                   ctrl.order_rules, ctrl._suggest_min_max)
                    self.bulk_tab._update_summary()
                self.status.showMessage(
                    f"Reverted {len(dialog.reverted_keys)} QOH edit(s)", 5000,
                )

    def _on_skip_actions(self):
        """Open skip cleanup tools dialog."""
        import skip_actions_flow
        model = self.bulk_tab.model if self.bulk_tab else None
        if model is None:
            return
        items = list(model.items)
        skip_items = skip_actions_flow.filter_skip_items(items)
        clusters = skip_actions_flow.count_skip_clusters_by_line_code(items)

        from ui_qt.workflow_dialogs import SkipActionsDialog
        dialog = SkipActionsDialog(skip_items, clusters, parent=self)

        def _on_action(action, scope):
            ctrl = self.controller
            if action == "ignore":
                keys = skip_actions_flow.collect_ignore_keys(scope)
                if keys:
                    ctrl.ignored_item_keys.update(keys)
                    try:
                        import storage
                        storage.save_ignored_items(
                            ctrl._data_path("ignored_items"), ctrl.ignored_item_keys,
                        )
                    except Exception:
                        pass
                    QMessageBox.information(dialog, "Ignored",
                                            f"Added {len(keys)} item(s) to ignore list.")
                    dialog.accept()
            elif action == "flag_discontinue":
                flagged = 0
                for lc, ic in skip_actions_flow.collect_keys_for_action(scope):
                    rule_key = f"{lc}:{ic}"
                    rule = ctrl.order_rules.setdefault(rule_key, {})
                    if not rule.get("discontinue_candidate"):
                        rule["discontinue_candidate"] = True
                        flagged += 1
                ctrl._save_order_rules()
                QMessageBox.information(dialog, "Flagged",
                                        f"Flagged {flagged} item(s) as discontinue candidates.")
            elif action == "export_csv":
                inv_lookup = ctrl.session.inventory_lookup if ctrl.session else {}
                rows = skip_actions_flow.build_skip_export_rows(scope, inv_lookup)
                if not rows:
                    QMessageBox.information(dialog, "Nothing to Export", "No usable rows.")
                    return
                path, _ = QFileDialog.getSaveFileName(
                    dialog, "Export Skip List", "skip_review.csv",
                    "CSV (*.csv);;All files (*.*)",
                )
                if not path:
                    return
                try:
                    csv_text = skip_actions_flow.render_skip_csv(rows)
                    with open(path, "w", encoding="utf-8", newline="") as fh:
                        fh.write(csv_text)
                    QMessageBox.information(dialog, "Exported",
                                            f"Wrote {len(rows)} row(s) to:\n{os.path.abspath(path)}")
                except OSError as exc:
                    QMessageBox.critical(dialog, "Export Failed", str(exc))

        dialog.action_requested.connect(_on_action)
        dialog.exec()

    def _on_session_history(self):
        """Open session history browser."""
        import storage
        ctrl = self.controller
        sessions_dir = ctrl._data_path("sessions")
        snapshots = storage.load_session_snapshots(sessions_dir, max_count=None)
        from ui_qt.workflow_dialogs import SessionHistoryDialog
        dialog = SessionHistoryDialog(snapshots, parent=self)
        dialog.exec()

    def _on_ignored_items(self):
        """Open ignored items manager."""
        ctrl = self.controller
        keys = sorted(ctrl.ignored_item_keys)
        from ui_qt.workflow_dialogs import IgnoredItemsDialog
        dialog = IgnoredItemsDialog(keys, parent=self)
        dialog.exec()
        if dialog.keys_to_remove:
            ctrl.ignored_item_keys -= dialog.keys_to_remove
            try:
                import storage
                storage.save_ignored_items(
                    ctrl._data_path("ignored_items"), ctrl.ignored_item_keys,
                )
            except Exception:
                pass
            self.status.showMessage(
                f"Removed {len(dialog.keys_to_remove)} item(s) from ignore list", 5000,
            )

    def _on_vendor_manager(self):
        """Open vendor manager dialog."""
        ctrl = self.controller
        codes = list(getattr(ctrl, "vendor_codes_used", None) or [])
        from ui_qt.workflow_dialogs import VendorManagerDialog
        dialog = VendorManagerDialog(codes, parent=self)
        dialog.exec()
        # Apply changes
        for change in dialog.changes:
            action = change.get("action")
            if action == "add":
                ctrl._remember_vendor_code(change["code"])
            elif action == "remove":
                ctrl._remove_vendor_code(change["code"])
            elif action == "rename":
                ctrl._rename_vendor_code(change["old"], change["new"])
