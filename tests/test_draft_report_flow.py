import os
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import draft_report_flow

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


def _item(lc, ic, vendor, final_qty, description="", qoh=0, cur_min=None, cur_max=None,
          pack_size=1, qty_on_po=0, why=""):
    return {
        "line_code": lc,
        "item_code": ic,
        "vendor": vendor,
        "description": description,
        "qoh": qoh,
        "cur_min": cur_min,
        "cur_max": cur_max,
        "pack_size": pack_size,
        "qty_on_po": qty_on_po,
        "final_qty": final_qty,
        "order_qty": final_qty,
        "why": why,
    }


class EligibleItemsTests(unittest.TestCase):
    def test_excludes_zero_qty(self):
        items = [
            _item("GR1-", "A", "GRELIN", final_qty=0),
            _item("GR1-", "B", "GRELIN", final_qty=5),
        ]
        out = draft_report_flow.eligible_items(items)
        self.assertEqual([i["item_code"] for i in out], ["B"])

    def test_excludes_unassigned(self):
        items = [
            _item("GR1-", "A", "", final_qty=5),
            _item("GR1-", "B", "   ", final_qty=5),
            _item("GR1-", "C", "GRELIN", final_qty=5),
        ]
        out = draft_report_flow.eligible_items(items)
        self.assertEqual([i["item_code"] for i in out], ["C"])


@unittest.skipIf(openpyxl is None, "openpyxl not installed")
class ExportDraftReviewFilesTests(unittest.TestCase):
    def _run(self, items, inventory_lookup=None, vendor_filter=None):
        inventory_lookup = inventory_lookup or {}
        with tempfile.TemporaryDirectory() as tmp:
            created = draft_report_flow.export_draft_review_files(
                items,
                inventory_lookup,
                tmp,
                run_date=datetime(2026, 4, 11, 10, 0, 0),
                vendor_filter=vendor_filter,
            )
            # Load each created workbook before the tempdir is torn down.
            snapshots = []
            for vendor, path in created:
                self.assertTrue(os.path.exists(path), f"missing output {path}")
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                snapshots.append((vendor, os.path.basename(path), [
                    [ws.cell(row=r, column=c).value for c in range(1, 12)]
                    for r in range(1, ws.max_row + 1)
                ], ws))
            return snapshots

    def test_writes_one_file_per_vendor(self):
        items = [
            _item("GR1-", "A", "GRELIN", final_qty=2, description="HOSE", pack_size=1),
            _item("GR1-", "B", "GRELIN", final_qty=3, description="BOLT", pack_size=1),
            _item("GR2-", "X", "MOTION", final_qty=1, description="BELT", pack_size=1),
        ]
        snapshots = self._run(items)
        self.assertEqual(len(snapshots), 2)
        vendors = [s[0] for s in snapshots]
        self.assertEqual(vendors, ["GRELIN", "MOTION"])
        self.assertTrue(snapshots[0][1].startswith("DraftReview_GRELIN_20260411"))

    def test_totals_row_sums_units_and_cost(self):
        items = [
            _item("GR1-", "A", "GRELIN", final_qty=2, description="HOSE"),
            _item("GR1-", "B", "GRELIN", final_qty=3, description="BOLT"),
        ]
        inv = {
            ("GR1-", "A"): {"qoh": 0, "repl_cost": 1.50, "min": 1, "max": 4},
            ("GR1-", "B"): {"qoh": 0, "repl_cost": 2.00, "min": 1, "max": 4},
        }
        snapshots = self._run(items, inventory_lookup=inv)
        _vendor, _fn, rows, _ws = snapshots[0]
        # Totals row contains TOTALS in col 1, total units in col 8, total ext in col 10.
        totals = [r for r in rows if r and r[0] == "TOTALS"]
        self.assertEqual(len(totals), 1)
        self.assertEqual(totals[0][7], 5)  # 2 + 3 units
        # 2*1.50 + 3*2.00 = 9.00
        self.assertAlmostEqual(totals[0][9], 9.0, places=2)

    def test_missing_repl_cost_leaves_ext_blank_but_still_counts_units(self):
        items = [
            _item("GR1-", "A", "GRELIN", final_qty=4, description="PART"),
        ]
        inv = {("GR1-", "A"): {"qoh": 0, "repl_cost": None, "min": 1, "max": 5}}
        snapshots = self._run(items, inventory_lookup=inv)
        _vendor, _fn, rows, _ws = snapshots[0]
        data_row = rows[3]  # row 4 = first data row (banner=1, subtitle=2, header=3)
        self.assertEqual(data_row[0], "A")
        self.assertEqual(data_row[7], 4)  # draft qty
        # Unit $ and Ext $ empty when cost unknown
        self.assertIn(data_row[8], ("", None))
        self.assertIn(data_row[9], ("", None))
        totals = [r for r in rows if r and r[0] == "TOTALS"]
        self.assertEqual(totals[0][7], 4)  # units still counted
        self.assertEqual(totals[0][9], 0)  # ext zero when no cost

    def test_vendor_filter_limits_output(self):
        items = [
            _item("GR1-", "A", "GRELIN", final_qty=2),
            _item("GR2-", "X", "MOTION", final_qty=1),
        ]
        snapshots = self._run(items, vendor_filter=["GRELIN"])
        self.assertEqual(len(snapshots), 1)
        self.assertEqual(snapshots[0][0], "GRELIN")

    def test_print_setup_flags_stamped(self):
        items = [_item("GR1-", "A", "GRELIN", final_qty=1)]
        snapshots = self._run(items)
        _vendor, _fn, _rows, ws = snapshots[0]
        # Landscape + fit-to-width stamped; header rows repeat.
        self.assertEqual(ws.page_setup.orientation, ws.ORIENTATION_LANDSCAPE)
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertTrue(ws.sheet_properties.pageSetUpPr.fitToPage)
        # openpyxl normalizes the range spec with $ anchors on reload.
        self.assertIn(ws.print_title_rows, ("1:3", "$1:$3"))

    def test_empty_items_creates_no_files(self):
        snapshots = self._run([])
        self.assertEqual(snapshots, [])

    def test_filename_sanitizes_vendor_slash(self):
        items = [_item("GR1-", "A", "BR/FARM", final_qty=1)]
        snapshots = self._run(items)
        self.assertEqual(len(snapshots), 1)
        filename = snapshots[0][1]
        # Forward slash must be sanitized out of filename
        self.assertNotIn("/", filename)
        self.assertTrue(filename.startswith("DraftReview_BR_FARM_"))


if __name__ == "__main__":
    unittest.main()
