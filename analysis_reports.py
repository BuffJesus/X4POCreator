"""Exportable analysis reports — dead stock, deferred items, session summary.

These reports give the operator actionable intelligence beyond the PO
files themselves.  Each function takes the session's filtered_items and
inventory_lookup and returns either structured data (for dialogs) or
writes a file (for export).

All functions are UI-agnostic — no Qt or tkinter imports.
"""

from __future__ import annotations

import csv
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Mapping, Sequence

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _vendor_or_unassigned(item: Mapping[str, Any]) -> str:
    return str(item.get("vendor", "") or "").strip().upper() or "(Unassigned)"


def _cost(item: Mapping[str, Any], inv_lookup: dict) -> float:
    cost = item.get("repl_cost")
    if not isinstance(cost, (int, float)):
        key = (item.get("line_code", ""), item.get("item_code", ""))
        cost = (inv_lookup.get(key) or {}).get("repl_cost")
    return float(cost) if isinstance(cost, (int, float)) else 0.0


# ── Dead Stock Report ────────────────────────────────────────────────

def build_dead_stock_rows(items: Sequence[dict], inv_lookup: dict) -> list[dict]:
    """Build export rows for all dead stock items, grouped by vendor."""
    rows = []
    for item in items:
        if not item.get("dead_stock"):
            continue
        inv = inv_lookup.get((item.get("line_code", ""), item.get("item_code", ""))) or {}
        unit_cost = _cost(item, inv_lookup)
        qoh = inv.get("qoh", 0) or 0
        rows.append({
            "vendor": _vendor_or_unassigned(item),
            "line_code": item.get("line_code", ""),
            "item_code": item.get("item_code", ""),
            "description": item.get("description", ""),
            "qoh": qoh,
            "unit_cost": unit_cost,
            "on_hand_value": round(unit_cost * qoh, 2) if unit_cost and qoh else 0,
            "last_sale": inv.get("last_sale", ""),
            "days_since_last_sale": item.get("days_since_last_sale", ""),
            "avg_days_between_sales": item.get("avg_days_between_sales", ""),
        })
    rows.sort(key=lambda r: (r["vendor"], r["line_code"], r["item_code"]))
    return rows


def dead_stock_summary(rows: list[dict]) -> dict:
    """Aggregate dead stock stats by vendor."""
    by_vendor: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_vendor[row["vendor"]].append(row)
    vendors = []
    total_value = 0.0
    for vendor in sorted(by_vendor):
        items = by_vendor[vendor]
        value = sum(r["on_hand_value"] for r in items)
        total_value += value
        vendors.append({
            "vendor": vendor,
            "item_count": len(items),
            "on_hand_value": round(value, 2),
        })
    return {
        "vendors": vendors,
        "total_items": len(rows),
        "total_on_hand_value": round(total_value, 2),
    }


def export_dead_stock_csv(items: Sequence[dict], inv_lookup: dict, output_dir: str) -> str:
    """Write a dead stock CSV and return the file path."""
    rows = build_dead_stock_rows(items, inv_lookup)
    timestamp = datetime.now().strftime("%Y%m%d")
    path = os.path.join(output_dir, f"DeadStock_{timestamp}.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Vendor", "Line Code", "Item Code", "Description",
            "QOH", "Unit Cost", "On Hand Value",
            "Last Sale", "Days Since Last Sale", "Avg Days Between Sales",
        ])
        for row in rows:
            writer.writerow([
                row["vendor"], row["line_code"], row["item_code"],
                row["description"], row["qoh"],
                f"{row['unit_cost']:.2f}" if row["unit_cost"] else "",
                f"{row['on_hand_value']:.2f}" if row["on_hand_value"] else "",
                row["last_sale"], row["days_since_last_sale"],
                row["avg_days_between_sales"],
            ])
    return path


def export_dead_stock_xlsx(items: Sequence[dict], inv_lookup: dict, output_dir: str) -> str:
    """Write a print-ready dead stock xlsx grouped by vendor. Returns file path."""
    rows = build_dead_stock_rows(items, inv_lookup)
    summary = dead_stock_summary(rows)
    timestamp = datetime.now().strftime("%Y%m%d")
    path = os.path.join(output_dir, f"DeadStock_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dead Stock"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    vendor_font = Font(bold=True, size=11, color="2C3E50")
    vendor_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    money_fmt = '#,##0.00'
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    totals_font = Font(bold=True, size=11)

    headers = [
        "Line Code", "Item Code", "Description", "QOH",
        "Unit Cost", "On Hand Value", "Last Sale",
        "Days Since Sale", "Avg Days Between Sales",
    ]
    col_widths = [10, 16, 30, 8, 12, 14, 12, 14, 18]

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    row_num = 1
    # Title
    ws.cell(row=row_num, column=1, value=f"Dead Stock Report — {datetime.now().strftime('%Y-%m-%d')}")
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    row_num += 1
    ws.cell(row=row_num, column=1,
            value=f"{summary['total_items']} items, ${summary['total_on_hand_value']:,.2f} total on-hand value")
    ws.cell(row=row_num, column=1).font = Font(size=10, italic=True)
    row_num += 2

    # Group by vendor
    by_vendor: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_vendor[r["vendor"]].append(r)

    for vendor in sorted(by_vendor):
        vitems = by_vendor[vendor]
        vendor_value = sum(r["on_hand_value"] for r in vitems)

        # Vendor header
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        vc = ws.cell(row=row_num, column=1,
                     value=f"{vendor}  —  {len(vitems)} items, ${vendor_value:,.2f}")
        vc.font = vendor_font
        vc.fill = vendor_fill
        row_num += 1

        # Column headers
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin
        row_num += 1

        # Data rows
        for r in vitems:
            vals = [
                r["line_code"], r["item_code"], r["description"], r["qoh"],
                r["unit_cost"] or None, r["on_hand_value"] or None,
                r["last_sale"], r["days_since_last_sale"],
                r["avg_days_between_sales"],
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.border = thin
                if ci in (5, 6):
                    c.number_format = money_fmt
                    c.alignment = Alignment(horizontal="right")
                if ci in (4, 8, 9):
                    c.alignment = Alignment(horizontal="center")
            row_num += 1

        # Vendor totals
        ws.cell(row=row_num, column=3, value="TOTAL").font = totals_font
        tc = ws.cell(row=row_num, column=6, value=vendor_value)
        tc.font = totals_font
        tc.number_format = money_fmt
        tc.border = thin
        row_num += 2

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = None

    wb.save(path)
    return path


# ── Deferred Items Report ────────────────────────────────────────────

def build_deferred_rows(items: Sequence[dict], inv_lookup: dict) -> list[dict]:
    """Build export rows for all pack-overshoot deferred items."""
    rows = []
    for item in items:
        if not item.get("deferred_pack_overshoot"):
            continue
        inv = inv_lookup.get((item.get("line_code", ""), item.get("item_code", ""))) or {}
        unit_cost = _cost(item, inv_lookup)
        qoh = inv.get("qoh", 0) or 0
        mx = inv.get("max", 0) or 0
        pct = int(round(qoh / mx * 100)) if mx else 0
        rows.append({
            "vendor": _vendor_or_unassigned(item),
            "line_code": item.get("line_code", ""),
            "item_code": item.get("item_code", ""),
            "description": item.get("description", ""),
            "qoh": qoh,
            "max": mx,
            "stock_pct": f"{pct}%",
            "pack_size": item.get("pack_size", ""),
            "raw_need": item.get("raw_need", 0),
            "unit_cost": unit_cost,
            "why": item.get("why", ""),
        })
    rows.sort(key=lambda r: (r["vendor"], r["line_code"], r["item_code"]))
    return rows


def export_deferred_csv(items: Sequence[dict], inv_lookup: dict, output_dir: str) -> str:
    """Write a deferred items CSV and return the file path."""
    rows = build_deferred_rows(items, inv_lookup)
    timestamp = datetime.now().strftime("%Y%m%d")
    path = os.path.join(output_dir, f"DeferredItems_{timestamp}.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Vendor", "Line Code", "Item Code", "Description",
            "QOH", "Max", "Stock %", "Pack Size", "Raw Need",
            "Unit Cost", "Why Deferred",
        ])
        for row in rows:
            writer.writerow([
                row["vendor"], row["line_code"], row["item_code"],
                row["description"], row["qoh"], row["max"],
                row["stock_pct"], row["pack_size"], row["raw_need"],
                f"{row['unit_cost']:.2f}" if row["unit_cost"] else "",
                row["why"],
            ])
    return path


# ── Session Summary Report ───────────────────────────────────────────

def build_session_summary(items: Sequence[dict], inv_lookup: dict) -> dict:
    """Build a comprehensive session summary with per-vendor breakdown."""
    by_vendor: dict[str, list[dict]] = defaultdict(list)
    total_items = 0
    assigned = 0
    unassigned = 0
    review_count = 0
    warning_count = 0
    skip_count = 0
    dead_stock_count = 0
    deferred_count = 0
    total_order_value = 0.0

    for item in items:
        total_items += 1
        vendor = _vendor_or_unassigned(item)
        status = str(item.get("status", "")).lower()
        has_vendor = bool(str(item.get("vendor", "")).strip())

        if has_vendor:
            assigned += 1
        else:
            unassigned += 1
        if status == "review":
            review_count += 1
        elif status in ("warning", "warn"):
            warning_count += 1
        elif status == "skip":
            skip_count += 1
        if item.get("dead_stock"):
            dead_stock_count += 1
        if item.get("deferred_pack_overshoot"):
            deferred_count += 1

        if has_vendor:
            qty = item.get("final_qty", 0) or 0
            cost = _cost(item, inv_lookup)
            value = qty * cost if isinstance(qty, (int, float)) and cost else 0
            total_order_value += value
            by_vendor[vendor].append(item)

    vendor_summaries = []
    for vendor in sorted(by_vendor):
        vitems = by_vendor[vendor]
        vvalue = sum(
            (it.get("final_qty", 0) or 0) * _cost(it, inv_lookup)
            for it in vitems
        )
        vendor_summaries.append({
            "vendor": vendor,
            "item_count": len(vitems),
            "order_value": round(vvalue, 2),
        })

    return {
        "total_items": total_items,
        "assigned": assigned,
        "unassigned": unassigned,
        "review_count": review_count,
        "warning_count": warning_count,
        "skip_count": skip_count,
        "dead_stock_count": dead_stock_count,
        "deferred_count": deferred_count,
        "total_order_value": round(total_order_value, 2),
        "vendor_summaries": vendor_summaries,
    }


def export_session_summary_csv(items: Sequence[dict], inv_lookup: dict, output_dir: str) -> str:
    """Write a session summary CSV and return the file path."""
    summary = build_session_summary(items, inv_lookup)
    timestamp = datetime.now().strftime("%Y%m%d")
    path = os.path.join(output_dir, f"SessionSummary_{timestamp}.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Session Summary", datetime.now().strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Items", summary["total_items"]])
        writer.writerow(["Assigned", summary["assigned"]])
        writer.writerow(["Unassigned", summary["unassigned"]])
        writer.writerow(["Needs Review", summary["review_count"]])
        writer.writerow(["Warnings", summary["warning_count"]])
        writer.writerow(["Skip (No Order)", summary["skip_count"]])
        writer.writerow(["Dead Stock", summary["dead_stock_count"]])
        writer.writerow(["Deferred (Pack Overshoot)", summary["deferred_count"]])
        writer.writerow(["Est. Total Order Value", f"${summary['total_order_value']:,.2f}"])
        writer.writerow([])
        writer.writerow(["Vendor", "Items", "Est. Order Value"])
        for vs in summary["vendor_summaries"]:
            writer.writerow([
                vs["vendor"], vs["item_count"],
                f"${vs['order_value']:,.2f}",
            ])
    return path
