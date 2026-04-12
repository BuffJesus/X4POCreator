"""Draft Review export — print-formatted per-vendor xlsx files.

The operator runs ``Draft Review`` from the bulk grid once vendor assignments
are roughed in.  For each vendor that has at least one item with
``order_qty > 0``, this module writes a landscape-letter xlsx sheet formatted
for physical print verification: item code, description, QOH, Min, Max, on-PO,
pack, draft qty, unit cost, extended cost, reason.  A bold totals row lists
the vendor's total units + total draft value.

Design decisions:
- One file per vendor (mirrors ``PO_VENDOR_YYYYMMDD.xlsx``) so the operator can
  print each vendor independently from File Explorer.
- Print settings are stamped at write time: landscape, fit-to-width=1,
  header row repeats on every page, page footer with vendor + page number.
- Unit cost pulls from ``shipping_flow.item_cost_data`` which already handles
  missing / zero / suspicious repl_cost gracefully; rows without a usable cost
  leave ext_cost blank (no zero-fill — that would lie to the operator).
- Items with ``order_qty <= 0`` are excluded (they have been triaged).
  Unassigned items are excluded (cannot be printed against a vendor that
  does not exist yet).
"""

from collections import defaultdict
from datetime import datetime
import os

import perf_trace
import shipping_flow

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - import guard
    raise ImportError("openpyxl is required for draft_report_flow") from exc


# Column layout — sized for landscape letter with ~0.5" margins.
# Widths are in Excel character units; tuned empirically for fit-to-width=1
# landscape printing without truncating the description column too far.
_COLUMNS = [
    ("item_code",   "Item Code",    16, "left"),
    ("description", "Description",  48, "left"),
    ("qoh",         "QOH",           7, "right"),
    ("cur_min",     "Min",           6, "right"),
    ("cur_max",     "Max",           6, "right"),
    ("qty_on_po",   "On PO",         7, "right"),
    ("pack_size",   "Pack",          6, "right"),
    ("draft_qty",   "Draft Qty",    10, "right"),
    ("unit_cost",   "Unit $",        9, "right"),
    ("ext_cost",    "Ext $",        11, "right"),
    # Blank columns for hand-written markup during physical verification
    ("_check",      "\u2713",        5, "center"),
    ("_adj_qty",    "Adj Qty",       9, "right"),
    ("_notes",      "Notes",        20, "left"),
]


def _safe_vendor_filename(vendor):
    safe = "".join(c if c.isalnum() or c in "-_ " else "_" for c in str(vendor or ""))
    return safe.strip() or "UNASSIGNED"


def _coerce_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_cost(value):
    if value is None:
        return ""
    return round(float(value), 2)


def _draft_qty(item):
    # Prefer final_qty (operator's commit number), fall back to order_qty.
    qty = item.get("final_qty")
    if qty in (None, ""):
        qty = item.get("order_qty", 0)
    try:
        return int(round(float(qty or 0)))
    except (TypeError, ValueError):
        return 0


def _format_why(item):
    why = item.get("why") or item.get("why_text") or ""
    # Bulk grid stores why text with embedded newlines and can be long; clip to
    # a reasonable length so it does not blow out the print row height.
    text = str(why).strip().replace("\r", " ").replace("\n", " ")
    if len(text) > 200:
        text = text[:197] + "..."
    return text


def _row_values(item, inventory_lookup, receipt_cost_lookup=None):
    key = (item.get("line_code", ""), item.get("item_code", ""))
    inv = (inventory_lookup or {}).get(key, {}) or {}
    draft_qty = _draft_qty(item)
    cost_data = shipping_flow.item_cost_data(
        {**item, "final_qty": draft_qty},
        inventory_lookup,
        receipt_cost_lookup=receipt_cost_lookup,
    )
    unit_cost = cost_data.get("unit_cost")
    cost_source = cost_data.get("source")
    has_cost = cost_source in ("inventory_repl_cost", "receipt_cost") and unit_cost is not None
    ext_cost = (unit_cost or 0.0) * draft_qty if has_cost else None

    # Prefer item's snapshot of min/max (already resolved by enrich); fall
    # back to inventory_lookup for the display-only review sheet.
    cur_min = item.get("cur_min")
    if cur_min in (None, ""):
        cur_min = inv.get("min")
    cur_max = item.get("cur_max")
    if cur_max in (None, ""):
        cur_max = inv.get("max")

    return {
        "item_code": item.get("item_code", ""),
        "description": item.get("description", "") or inv.get("description", ""),
        "qoh": _coerce_number(item.get("qoh", inv.get("qoh"))),
        "cur_min": _coerce_number(cur_min),
        "cur_max": _coerce_number(cur_max),
        "qty_on_po": _coerce_number(item.get("qty_on_po", 0)) or 0,
        "pack_size": _coerce_number(item.get("pack_size")),
        "draft_qty": draft_qty,
        "unit_cost": _format_cost(unit_cost) if has_cost else "",
        "ext_cost": _format_cost(ext_cost) if ext_cost is not None else "",
        "_ext_cost_numeric": ext_cost or 0.0,
        "why": _format_why(item),
    }


def _sort_key(row):
    return (str(row.get("item_code", "")),)


def eligible_items(items):
    """Filter items to those worth printing for verification.

    Excludes zeroed-out rows and unassigned rows.  Public so tests and
    callers that want to preview the filtered set don't need to duplicate
    the logic.
    """
    out = []
    for item in items or []:
        if not str(item.get("vendor", "") or "").strip():
            continue
        if _draft_qty(item) <= 0:
            continue
        out.append(item)
    return out


def group_by_vendor(items):
    groups = defaultdict(list)
    for item in items:
        vendor = str(item.get("vendor", "") or "").strip() or "UNASSIGNED"
        groups[vendor].append(item)
    return groups


def _apply_print_setup(ws, vendor, total_pages_hint=None):
    """Stamp the sheet with print settings targeting landscape letter."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # unlimited pages vertically
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    # Repeat header row on every printed page (rows 1-3: banner + column header).
    # openpyxl expects a range string like "1:3".
    ws.print_title_rows = "1:3"
    ws.oddHeader.center.text = f"Draft PO Review — {vendor}"
    ws.oddHeader.center.size = 11
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddFooter.left.text = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.oddFooter.center.text = f"{vendor}"
    ws.oddFooter.right.text = "Page &P of &N"


def _write_vendor_sheet(wb, vendor, items, inventory_lookup, *, run_date, receipt_cost_lookup=None):
    ws = wb.active if wb.worksheets and wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = (vendor[:28] or "DRAFT").replace("/", "_")

    # ── Design tokens (Tuner-inspired) ────────────────────────────
    # Dark header with high contrast, warm accent on the draft qty
    # column, alternating row tints for scannability, generous row
    # height so the operator never squints.
    _DEEP = "14171E"       # near-black header background
    _PANEL = "1A1D24"      # dark panel for totals
    _ROW_EVEN = "F8F9FA"   # light grey even rows (print-friendly)
    _ROW_ODD = "FFFFFF"    # white odd rows
    _ACCENT = "5A9AD6"     # blue accent (matches theme.ACCENT_PRIMARY)
    _DRAFT_BG = "FFF8E1"   # warm yellow for draft qty column
    _DRAFT_BOLD = "F9A825"  # amber for draft qty header
    _OK = "43A047"         # green for totals
    _TEXT_W = "FFFFFF"      # white text on dark
    _TEXT_D = "212121"      # dark text on light
    _TEXT_M = "757575"      # muted text
    _BORDER = "E0E0E0"     # soft border for rows

    banner_font = Font(bold=True, size=14, color=_TEXT_D)
    date_font = Font(size=10, color=_TEXT_M)
    sub_font = Font(size=10, italic=True, color=_TEXT_M)
    header_font = Font(bold=True, size=9, color=_TEXT_W)
    header_fill = PatternFill(start_color=_DEEP, end_color=_DEEP, fill_type="solid")
    draft_header_fill = PatternFill(start_color=_DRAFT_BOLD, end_color=_DRAFT_BOLD, fill_type="solid")
    body_font = Font(size=9, color=_TEXT_D)
    draft_qty_font = Font(bold=True, size=10, color=_TEXT_D)
    draft_qty_fill = PatternFill(start_color=_DRAFT_BG, end_color=_DRAFT_BG, fill_type="solid")
    cost_est_font = Font(size=9, italic=True, color=_TEXT_M)
    totals_font = Font(bold=True, size=11, color=_TEXT_W)
    totals_fill = PatternFill(start_color=_DEEP, end_color=_DEEP, fill_type="solid")
    totals_draft_fill = PatternFill(start_color=_OK, end_color=_OK, fill_type="solid")
    even_fill = PatternFill(start_color=_ROW_EVEN, end_color=_ROW_EVEN, fill_type="solid")
    soft = Side(style="thin", color=_BORDER)
    border = Border(left=soft, right=soft, top=soft, bottom=soft)
    no_border = Border()

    n_cols = len(_COLUMNS)

    # Row 1: banner — vendor name prominently
    banner_cell = ws.cell(row=1, column=1, value=f"Draft PO Review  -  {vendor}")
    banner_cell.font = banner_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols - 3))
    date_cell = ws.cell(row=1, column=n_cols - 2, value=run_date.strftime("%A, %B %d, %Y"))
    date_cell.font = date_font
    date_cell.alignment = Alignment(horizontal="right")
    ws.merge_cells(start_row=1, start_column=n_cols - 2, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 28

    # Row 2: subtitle — item count + instructions
    subtitle = f"{len(items)} items  |  Yellow column = draft order qty"
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 20

    # Row 3: column headers — dark background, draft qty in amber,
    # markup columns in light grey to signal "write here"
    markup_fill = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
    markup_header_font = Font(bold=True, size=9, color=_TEXT_M)
    header_row = 3
    for idx, (key, label, _width, _align) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        if key.startswith("_"):
            cell.font = markup_header_font
            cell.fill = markup_fill
        elif key == "draft_qty":
            cell.font = header_font
            cell.fill = draft_header_fill
        else:
            cell.font = header_font
            cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 24

    # Data rows — alternating tints, generous height
    rows_data = [_row_values(item, inventory_lookup, receipt_cost_lookup) for item in items]
    rows_data.sort(key=_sort_key)

    total_units = 0
    total_ext = 0.0
    items_with_cost = 0
    for row_offset, row_data in enumerate(rows_data, start=header_row + 1):
        is_even = (row_offset - header_row) % 2 == 0
        for col_idx, (key, _label, _width, align) in enumerate(_COLUMNS, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.border = border
            cell.font = body_font
            cell.alignment = Alignment(
                horizontal=align,
                vertical="top",
                wrap_text=(key in ("description", "why")),
            )
            # Draft qty column: always warm yellow
            if key == "draft_qty":
                cell.font = draft_qty_font
                cell.fill = draft_qty_fill
            elif is_even:
                cell.fill = even_fill
            # Cost columns: number format + italic if receipt-based
            if key in ("unit_cost", "ext_cost") and isinstance(value, (int, float)):
                cell.number_format = "$#,##0.00"
        ws.row_dimensions[row_offset].height = 18
        total_units += int(row_data.get("draft_qty", 0) or 0)
        ext = row_data.get("_ext_cost_numeric", 0.0) or 0.0
        total_ext += ext
        if ext > 0:
            items_with_cost += 1

    # Totals row — dark background, green draft qty, prominent
    totals_row = header_row + len(rows_data) + 2  # blank row before totals
    ws.row_dimensions[totals_row - 1].height = 8  # spacer
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=totals_row, column=col)
        cell.fill = totals_fill
        cell.border = border

    label_cell = ws.cell(row=totals_row, column=1, value="TOTALS")
    label_cell.font = totals_font
    label_cell.fill = totals_fill
    ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=6)

    draft_col = 8
    ext_col = 10
    why_col = 11

    units_cell = ws.cell(row=totals_row, column=draft_col, value=total_units)
    units_cell.font = Font(bold=True, size=12, color=_TEXT_W)
    units_cell.fill = totals_draft_fill
    units_cell.alignment = Alignment(horizontal="right")
    units_cell.border = border

    ext_cell = ws.cell(row=totals_row, column=ext_col, value=round(total_ext, 2))
    ext_cell.font = totals_font
    ext_cell.alignment = Alignment(horizontal="right")
    ext_cell.number_format = "$#,##0.00"
    ext_cell.border = border

    # Fill remaining totals row cells with the dark background
    for col in range(why_col, n_cols + 1):
        c = ws.cell(row=totals_row, column=col)
        c.fill = totals_fill
        c.border = border

    ws.row_dimensions[totals_row].height = 26

    # Column widths
    for idx, (_key, _label, width, _align) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    _apply_print_setup(ws, vendor)
    return ws


@perf_trace.timed("draft_report_flow.export_draft_review_files")
def export_draft_review_files(
    items,
    inventory_lookup,
    output_dir,
    *,
    run_date=None,
    vendor_filter=None,
    receipt_cost_lookup=None,
):
    """Write one xlsx per vendor; return list of (vendor, path) tuples.

    Parameters
    ----------
    items : iterable of bulk-grid item dicts (``filtered_items``).
    inventory_lookup : session.inventory_lookup (for unit cost and min/max fallback).
    output_dir : directory the files are written into.  Must exist.
    run_date : optional datetime for the banner/footer; defaults to now.
    vendor_filter : optional iterable of vendor codes.  When set, only those
        vendors' files are written.
    """
    if run_date is None:
        run_date = datetime.now()
    if vendor_filter is not None:
        vendor_filter = {str(v or "").strip() for v in vendor_filter if v}

    eligible = eligible_items(items)
    if vendor_filter:
        eligible = [i for i in eligible if str(i.get("vendor", "")).strip() in vendor_filter]
    vendor_groups = group_by_vendor(eligible)

    created = []
    timestamp = run_date.strftime("%Y%m%d")
    for vendor, vitems in sorted(vendor_groups.items()):
        vitems_sorted = sorted(vitems, key=lambda it: str(it.get("item_code", "")))
        wb = openpyxl.Workbook()
        _write_vendor_sheet(wb, vendor, vitems_sorted, inventory_lookup,
                           run_date=run_date, receipt_cost_lookup=receipt_cost_lookup)
        safe_vendor = _safe_vendor_filename(vendor)
        filename = f"DraftReview_{safe_vendor}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        created.append((vendor, filepath))
    return created
