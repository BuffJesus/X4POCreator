#!/usr/bin/env python3
"""
PO Builder - Purchase Order generation tool for X4 Import from Excel.

Reads Part Sales & Receipts CSVs (and optionally suspended items / open PO listings),
lets the user exclude line codes, assign vendor codes item-by-item, review/edit
quantities, and exports one .xlsx file per vendor in the X4 import format.
"""

import csv
import os
import sys
import threading
import re
import copy
import urllib.request
import urllib.error
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from datetime import datetime
import json
from app_version import APP_VERSION as INTERNAL_APP_VERSION
import app_runtime_flow
import assignment_flow
import bulk_context_flow
import bulk_edit_flow
import bulk_sheet_actions_flow
import data_folder_flow
import export_flow
import item_workflow
import load_flow
import loading_flow
import maintenance_flow
import parsers
import persistent_state_flow
import reorder_flow
import review_flow
import session_state_flow
import shipping_flow
import storage
import ui_state_flow
from debug_log import DEBUG_LOG_FILE, write_debug
from bulk_sheet import BulkSheetView
import ui_assignment_actions
import ui_bulk
import ui_bulk_dialogs
import ui_filters
import ui_help
import ui_individual
import ui_load
import ui_review
import ui_vendor_manager
from maintenance import build_maintenance_report
from models import AppSessionState, ItemKey, MaintenanceCandidate, SessionItemState, SourceItemState, SuggestedItemState
from rules import (
    enrich_item,
    evaluate_item_status,
)

try:
    import winsound
    HAS_WINSOUND = True
except ImportError:
    HAS_WINSOUND = False

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# Persistent files stored next to the script
# Path setup — works both as a script and as a PyInstaller .exe
# Bundled assets (loading.gif) go to the PyInstaller temp dir
# User data (whitelist, history) stays next to the executable so it persists
if getattr(sys, "frozen", False):
    # Running as PyInstaller bundle
    _BUNDLE_DIR = sys._MEIPASS           # temp dir with bundled assets
    _DATA_DIR = os.path.dirname(sys.executable)  # folder containing the .exe
else:
    _BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    _DATA_DIR = _BUNDLE_DIR

LOCAL_DATA_DIR = _DATA_DIR
APP_SETTINGS_FILE = os.path.join(_DATA_DIR, "po_builder_settings.json")
VERSION_FILE = os.path.join(_DATA_DIR, "VERSION")
DUPLICATE_WHITELIST_FILE = os.path.join(LOCAL_DATA_DIR, "duplicate_whitelist.txt")
ORDER_HISTORY_FILE = os.path.join(LOCAL_DATA_DIR, "order_history.json")
ORDER_RULES_FILE = os.path.join(LOCAL_DATA_DIR, "order_rules.json")
SUSPENSE_CARRY_FILE = os.path.join(LOCAL_DATA_DIR, "suspense_carry.json")
SESSIONS_DIR = os.path.join(LOCAL_DATA_DIR, "sessions")
VENDOR_CODES_FILE = os.path.join(LOCAL_DATA_DIR, "vendor_codes.txt")
IGNORED_ITEMS_FILE = os.path.join(LOCAL_DATA_DIR, "ignored_items.txt")
GITHUB_REPO = "BuffJesus/X4POCreator"
GITHUB_RELEASES_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
GITHUB_RELEASES_PAGE_URL = f"https://github.com/{GITHUB_REPO}/releases/latest"
LOADING_GIF_FILE = os.path.join(_BUNDLE_DIR, "loading.gif")
LOADING_WAV_FILE = os.path.join(_BUNDLE_DIR, "loading.wav")
ICON_FILE = os.path.join(_BUNDLE_DIR, "icon.ico")
OLD_PO_WARNING_DAYS = 90
SHORT_SALES_WINDOW_DAYS = 7
MIN_ANNUAL_SALES_FOR_SUGGESTIONS = 3
MAX_EXCEED_MARGIN = 1.5
MAX_EXCEED_ABS_BUFFER = 5
MAX_LOADING_GIF_FRAMES = 90
CORNER_LOADING_GIF_SIZE = (52, 52)
BULK_EDITABLE_COLS = ("vendor", "final_qty", "qoh", "cur_min", "cur_max", "pack_size")
REVIEW_EDITABLE_COLS = ("vendor", "order_qty", "pack_size")
DEFAULT_MIXED_EXPORT_BEHAVIOR = "all_exportable"
MIXED_EXPORT_BEHAVIOR_OPTIONS = ("all_exportable", "immediate_only", "ask_when_mixed")
DEFAULT_PLANNED_ONLY_EXPORT_BEHAVIOR = "export_automatically"
PLANNED_ONLY_EXPORT_BEHAVIOR_OPTIONS = ("export_automatically", "ask_before_export")
DEFAULT_REVIEW_EXPORT_FOCUS = "exceptions_only"
REVIEW_EXPORT_FOCUS_OPTIONS = ("all_items", "exceptions_only")
DEFAULT_REMOVE_NOT_NEEDED_SCOPE = "unassigned_only"
REMOVE_NOT_NEEDED_SCOPE_OPTIONS = ("unassigned_only", "include_assigned")
DEFAULT_VENDOR_POLICY_PRESET = ""
BULK_SHORTCUTS_TEXT = """Current bulk-sheet shortcuts

Supported now
- Ctrl+C: copy selected cells or rows
- Ctrl+V: paste into the active editable area
- Ctrl+A: select all visible rows
- Ctrl+Z: undo the last bulk edit or row removal
- Ctrl+Y: redo the last undone bulk edit or row removal
- Ctrl+D: fill down using the current cell value
- Ctrl+R: same as Ctrl+D for the current one-column bulk edit model
- Ctrl+Enter: apply the current cell value across the selected rows in the active editable column
- Delete / Backspace: clear selected cells, or remove selected rows
- F2 or Enter: edit the current editable selection
- Tab / Shift+Tab: move the active cell across editable bulk columns
- Shift+Arrow: extend the current cell selection range
- Home / End: jump to first or last editable column on the current row
- Ctrl+Arrow: jump to row or editable-column edges in the current direction
- Esc: clear the current selection
- Shift+Space: select current row
- Ctrl+Space: select current column

Planned next
- commit-and-move behavior while actively editing cells
- Page Up / Page Down: stronger spreadsheet navigation
"""
MAX_BULK_HISTORY = 25


def _session_field(name):
    def _get(self):
        return getattr(self.session, name)

    def _set(self, value):
        setattr(self.session, name, value)
        if name == "filtered_items":
            ui_bulk.sync_bulk_cache_state(self, filtered_items_changed=True, retain_items=value)
            ui_bulk.sync_bulk_session_metadata(self, value)

    return property(_get, _set)


# ─── Order Rules (persistent per-item buy rules) ────────────────────────────

def get_rule_key(line_code, item_code):
    """Build a consistent key for the order rules dict."""
    return f"{line_code}:{item_code}"


def load_app_version(path=VERSION_FILE, default=INTERNAL_APP_VERSION):
    candidate_paths = [path]
    bundled_path = os.path.join(_BUNDLE_DIR, "VERSION")
    if bundled_path not in candidate_paths:
        candidate_paths.append(bundled_path)
    for candidate in candidate_paths:
        try:
            with open(candidate, "r", encoding="utf-8") as handle:
                value = handle.read().strip()
                if value:
                    return value
        except Exception:
            continue
    return default


APP_VERSION = load_app_version()


def _parse_version_parts(value):
    normalized = str(value or "").strip().lstrip("vV")
    match = re.match(r"^(\d+)\.(\d+)\.(\d+)$", normalized)
    if not match:
        return None
    return tuple(int(part) for part in match.groups())


def is_release_version(value):
    return _parse_version_parts(value) is not None


def is_newer_version(candidate, current):
    candidate_parts = _parse_version_parts(candidate)
    current_parts = _parse_version_parts(current)
    if candidate_parts is None or current_parts is None:
        return False
    return candidate_parts > current_parts


def fetch_latest_github_release(url=GITHUB_RELEASES_API_URL, timeout=3):
    request = urllib.request.Request(
        url,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": f"PO-Builder/{APP_VERSION}",
        },
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return {
        "tag_name": str(payload.get("tag_name", "")).strip(),
        "html_url": str(payload.get("html_url", "")).strip() or GITHUB_RELEASES_PAGE_URL,
        "name": str(payload.get("name", "")).strip(),
        "published_at": str(payload.get("published_at", "")).strip(),
    }

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ─── Known Vendor Codes ─────────────────────────────────────────────────────

KNOWN_VENDORS = [
    "18.80", "ABOVBEYO", "ABV&BEY", "ACKLANDS", "ACKLGRU", "ACKLGRUN",
    "ACT AIR", "ACTIVE", "ADVWARE", "ALBECHAI", "ALL-TEK", "ALLIDIBR",
    "ALLIDICA", "ALLIED", "AMAZON", "AMSOIL", "ANDEFRED", "APPIND", "APPLID",
    "APPLIED", "ARMALL", "ATLATRAI", "ATS ELEC", "B/B", "BAT DIR", "BATDIR",
    "BATT/DIR", "BATTDIRE", "BCBEAR", "BEARTRAN", "BEREFLUI", "BOB DALE",
    "BOBDALE", "BOLTSU", "BOLTSUP", "BOLTSUPP", "BOMAENTE", "BR/FARM",
    "BRICON", "BROOINDU", "BROOKS F",
    # Image 2
    "CA/PUMP", "CAL-SCAN", "CAMPBELL", "CANAPUMP", "CANYON", "CANYRIGG",
    "CASCADE", "CASLAN", "CELDIS", "CHRIPAGE", "COMMERCI", "COMMSOLU",
    "COSTCO", "CROSCANA", "CROSS", "CROSSCA", "CROSSCAN", "CTP",
    "DAYCCANA", "DAYCO", "DEFIOPTI", "DENDOFF", "DERKMANU", "DIX PERF",
    "DOCACORP", "DOLLAR", "DRIVE PR", "DRIVPROD", "DYL", "DYNAINDU",
    "DYNALINE", "EDMONTON", "EDMONUT", "ELEMETDE", "FISHSCIE", "FLUID",
    "FLUISEAL", "FORNEY", "FORNWELD", "FOXSAFE", "FREDANDE",
    # Image 3
    "FTGAINDU", "G2S", "G2SEQUIP", "GARAPAK", "GAZEOILF", "GEARCENT",
    "GHJULIMI", "GLOBAL", "GLOBAUT", "GOODRUBB", "GR/WEST", "GRANENTE",
    "GREDIST", "GREEN", "GREENLI", "GREGDIST", "GREGG", "GREGGS", "GRELIN",
    "GRGGS", "HARWOOD", "HITAKOKI", "HJU", "HOME DEP", "HOME HAR", "HOMEHA",
    "HPAULTUN", "HYDRSTEE", "HYDRX", "IDENT", "IFRWORK", "IND ENG",
    "INDENG", "J-DON", "JASOINDU", "JASON", "JBROS", "JET", "JETEQUI",
    "JOHNBROO", "JONELL",
    # Image 4
    "KDS", "KDSMUR", "KENCOVE", "KINEINCO", "KOYOCANA", "LEEVALL", "LJPETE",
    "MACMINDU", "MAGIAUTO", "MAINFILT", "MBL", "MCCOBROS", "MEDIBATT",
    "MIDFSUPP", "MILLER", "MILSUP", "MISC", "MOTION", "MRC", "NACHCANA",
    "NACHI", "NAPDIS", "NO SPILL", "NORDA-TE", "NORDTECH", "NORFLU",
    "NORTMETT", "NORWINDU", "NOSPILL", "P.S.I", "PADDPLAS", "PANAMA",
    "PAPCO", "PARA", "PARAHYD", "PARAHYDR", "PARAMECH", "PARAMNT", "PARHYDR",
    "PATS", "PATSDRIV",
    # Image 5
    "PECOFACE", "PETROCAN", "PIPETECH", "POWEIGNI", "POWER", "PREMIUM",
    "PREMTOOL", "PROLINE", "PTA", "PTMIND", "PWR IGN", "R&M", "RAYNAUTO",
    "RBW", "RBWIL", "RBWILL", "RE XSTRM", "REDLDIST", "RELIINDU",
    "REMA TIP", "RENOINDU", "ROBETHIB", "ROLG", "SEALWELD", "SHELMACH",
    "SHIPSUPP", "SHURSEAL", "SKEANS", "SOURCE", "SPAENAUR", "SPALHARD",
    "SPARTAN", "STAPEWAY", "STAPLES", "STEERACI", "SUMMIT", "SWAGELOK",
    "SYSTPLUS", "THERTECH", "TITASUPP", "TRANEXPR",
    # Image 6
    "TRANSUPP", "TSL", "UKPROD", "UNI", "UNI/DYL", "UNI/WIX", "UNIFIED",
    "UNISE", "UNISEL", "UNISEL;", "UNISELE", "UNIVALV", "UNSIELE", "WAJIND",
    "WAJIND01", "WAL/COST", "WALLMACH", "WALMART", "WATSGLOV", "WATSON",
    "WCSC", "WESTERN", "WESTGAUG", "WESTINDU", "WESTPART", "WESTWARD", "WIL",
    "WILSAUTO", "WURTCANA", "WURTH",
]


# ─── Excel Export ─────────────────────────────────────────────────────────────

def export_vendor_po(vendor_code, items, output_dir):
    """
    Export a single vendor PO as an .xlsx file in X4 import format.
    Headers: product group, item code, order qty
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Import"

    # Header row styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["product group", "item code", "order quantity"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, item in enumerate(items, 2):
        pg_cell = ws.cell(row=row_idx, column=1, value=item["line_code"])
        ic_cell = ws.cell(row=row_idx, column=2, value=item["item_code"])
        qty_cell = ws.cell(row=row_idx, column=3, value=item["order_qty"])
        for cell in (pg_cell, ic_cell, qty_cell):
            cell.border = thin_border
        # Force item code as text to preserve leading zeros
        ic_cell.number_format = "@"
        qty_cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in vendor_code)
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"PO_{safe_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ─── GUI Application ─────────────────────────────────────────────────────────

class POBuilderApp:
    sales_items = _session_field("sales_items")
    po_items = _session_field("po_items")
    suspended_items = _session_field("suspended_items")
    suspended_set = _session_field("suspended_set")
    suspended_lookup = _session_field("suspended_lookup")
    open_po_lookup = _session_field("open_po_lookup")
    all_line_codes = _session_field("all_line_codes")
    inventory_lookup = _session_field("inventory_lookup")
    inventory_source_lookup = _session_field("inventory_source_lookup")
    pack_size_lookup = _session_field("pack_size_lookup")
    pack_size_source_lookup = _session_field("pack_size_source_lookup")
    pack_size_by_item = _session_field("pack_size_by_item")
    pack_size_conflicts = _session_field("pack_size_conflicts")
    on_po_qty = _session_field("on_po_qty")
    qoh_adjustments = _session_field("qoh_adjustments")
    duplicate_ic_lookup = _session_field("duplicate_ic_lookup")
    recent_orders = _session_field("recent_orders")
    vendor_policies = _session_field("vendor_policies")
    order_rules = _session_field("order_rules")
    suspense_carry = _session_field("suspense_carry")
    vendor_codes_used = _session_field("vendor_codes_used")
    filtered_items = _session_field("filtered_items")
    individual_items = _session_field("individual_items")
    assigned_items = _session_field("assigned_items")
    startup_warning_rows = _session_field("startup_warning_rows")
    def __init__(self, root):
        self.root = root
        self.app_settings = self._load_app_settings()
        self._startup_data_dir_warning = ""
        self.shared_data_dir = ""
        self.data_dir = LOCAL_DATA_DIR
        self.data_paths = {}
        self.update_check_enabled = False
        self.session = AppSessionState()
        self._configure_initial_data_dir()
        self.root.title("PO Builder — X4 Import Tool")
        self.root.geometry("1100x720")
        self.root.minsize(900, 600)

        # State
        self.excluded_line_codes = set()
        self.all_customers = []         # (code, name, count) from suspended items
        self.excluded_customers = set() # customer codes to exclude
        self.dup_whitelist = set()      # persistent whitelist
        self.ignored_item_keys = set()  # persistent ignore list keyed by LC:Item
        self.last_removed_bulk_items = []  # [(index, item_dict)] for one-step undo
        self.bulk_undo_stack = []
        self.bulk_redo_stack = []
        self.bulk_sheet = None
        self.review_grid_editor = None
        self._loaded_dup_whitelist = set()
        self._loaded_ignored_item_keys = set()
        self._loaded_order_rules = {}
        self._loaded_suspense_carry = {}
        self._loaded_vendor_codes = []
        self._load_persistent_state()

        # Build the notebook (tabbed interface) — styles set by apply_dark_theme()
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Loading overlay
        self._loading_overlay = None
        self._loading_frames = []
        self._loading_frame_idx = 0
        self._loading_after_id = None
        self._loading_frames_loaded = False
        self._corner_loading_label = None
        self._corner_loading_frames = []
        self._corner_loading_frame_idx = 0
        self._corner_loading_after_id = None
        self._corner_loading_frames_loaded = False

        self._build_tab_load()
        self._build_tab_exclude()
        self._build_tab_customers()
        self._build_tab_bulk_assign()
        self._build_tab_individual_assign()
        self._build_tab_review()
        self._build_tab_help()
        self._show_corner_loading_gif()

        # Disable tabs 2-6 until data is loaded
        for i in (1, 2, 3, 4, 5):
            self.notebook.tab(i, state="disabled")
        if self._startup_data_dir_warning:
            self.root.after(100, lambda: messagebox.showwarning("Shared Data Folder", self._startup_data_dir_warning))
        if self.update_check_enabled:
            self.root.after(1500, self._start_update_check)

    # ── Loading Overlay ──────────────────────────────────────────────────

    @staticmethod
    def _build_data_paths(data_dir):
        return data_folder_flow.build_data_paths(data_dir)

    def _data_path(self, key):
        return self.data_paths[key]

    def _load_app_settings(self):
        settings = storage.load_json_file(APP_SETTINGS_FILE, {})
        return settings if isinstance(settings, dict) else {}

    def _save_app_settings(self):
        app_runtime_flow.save_app_settings(self, APP_SETTINGS_FILE, write_debug)

    def _configure_initial_data_dir(self):
        data_folder_flow.configure_initial_data_dir(self)

    def _load_persistent_state(self):
        data_folder_flow.load_persistent_state(self, KNOWN_VENDORS)

    def _active_data_folder_label(self):
        return data_folder_flow.active_data_folder_label(self)

    def _refresh_data_folder_labels(self):
        data_folder_flow.refresh_data_folder_labels(self)

    def _open_active_data_folder(self):
        app_runtime_flow.open_active_data_folder(self)

    def _rebuild_duplicate_ic_lookup(self):
        data_folder_flow.rebuild_duplicate_ic_lookup(self)

    def _prune_ignored_items_from_session(self):
        return data_folder_flow.prune_ignored_items_from_session(self)

    def _has_active_assignment_session(self):
        return data_folder_flow.has_active_assignment_session(self)

    def _refresh_active_data_state(self, notify=True):
        return data_folder_flow.refresh_active_data_state(self, KNOWN_VENDORS, get_rule_key, notify=notify)

    def _set_shared_data_folder(self):
        data_folder_flow.set_shared_data_folder(self, KNOWN_VENDORS, get_rule_key)

    def _use_local_data_folder(self):
        data_folder_flow.use_local_data_folder(self, LOCAL_DATA_DIR, KNOWN_VENDORS, get_rule_key)

    def _set_update_check_enabled(self):
        app_runtime_flow.set_update_check_enabled(self)

    def _get_mixed_export_behavior(self):
        behavior = str(self.app_settings.get("mixed_export_behavior", DEFAULT_MIXED_EXPORT_BEHAVIOR) or "").strip()
        if behavior not in MIXED_EXPORT_BEHAVIOR_OPTIONS:
            behavior = DEFAULT_MIXED_EXPORT_BEHAVIOR
        return behavior

    def _set_mixed_export_behavior(self, behavior):
        normalized = str(behavior or "").strip()
        if normalized not in MIXED_EXPORT_BEHAVIOR_OPTIONS:
            normalized = DEFAULT_MIXED_EXPORT_BEHAVIOR
        self.app_settings["mixed_export_behavior"] = normalized
        self._save_app_settings()

    def _get_review_export_focus(self):
        focus = str(self.app_settings.get("review_export_focus", DEFAULT_REVIEW_EXPORT_FOCUS) or "").strip()
        if focus not in REVIEW_EXPORT_FOCUS_OPTIONS:
            focus = DEFAULT_REVIEW_EXPORT_FOCUS
        return focus

    def _set_review_export_focus(self, focus):
        normalized = str(focus or "").strip()
        if normalized not in REVIEW_EXPORT_FOCUS_OPTIONS:
            normalized = DEFAULT_REVIEW_EXPORT_FOCUS
        self.app_settings["review_export_focus"] = normalized
        self._save_app_settings()

    def _get_planned_only_export_behavior(self):
        behavior = str(self.app_settings.get("planned_only_export_behavior", DEFAULT_PLANNED_ONLY_EXPORT_BEHAVIOR) or "").strip()
        if behavior not in PLANNED_ONLY_EXPORT_BEHAVIOR_OPTIONS:
            behavior = DEFAULT_PLANNED_ONLY_EXPORT_BEHAVIOR
        return behavior

    def _set_planned_only_export_behavior(self, behavior):
        normalized = str(behavior or "").strip()
        if normalized not in PLANNED_ONLY_EXPORT_BEHAVIOR_OPTIONS:
            normalized = DEFAULT_PLANNED_ONLY_EXPORT_BEHAVIOR
        self.app_settings["planned_only_export_behavior"] = normalized
        self._save_app_settings()

    def _get_remove_not_needed_scope(self):
        scope = str(self.app_settings.get("remove_not_needed_scope", DEFAULT_REMOVE_NOT_NEEDED_SCOPE) or "").strip()
        if scope not in REMOVE_NOT_NEEDED_SCOPE_OPTIONS:
            scope = DEFAULT_REMOVE_NOT_NEEDED_SCOPE
        return scope

    def _set_remove_not_needed_scope(self, scope):
        normalized = str(scope or "").strip()
        if normalized not in REMOVE_NOT_NEEDED_SCOPE_OPTIONS:
            normalized = DEFAULT_REMOVE_NOT_NEEDED_SCOPE
        self.app_settings["remove_not_needed_scope"] = normalized
        self._save_app_settings()

    def _get_last_export_dir(self):
        value = str(self.app_settings.get("last_export_dir", "") or "").strip()
        return value

    def _set_last_export_dir(self, path):
        normalized = str(path or "").strip()
        self.app_settings["last_export_dir"] = normalized
        self._save_app_settings()

    def _get_default_vendor_policy_preset(self):
        preset_name = str(self.app_settings.get("default_vendor_policy_preset", DEFAULT_VENDOR_POLICY_PRESET) or "").strip()
        valid_names = {key for key, _label in shipping_flow.vendor_policy_preset_options()}
        if preset_name not in valid_names:
            preset_name = DEFAULT_VENDOR_POLICY_PRESET
        return preset_name

    def _set_default_vendor_policy_preset(self, preset_name):
        normalized = str(preset_name or "").strip()
        valid_names = {key for key, _label in shipping_flow.vendor_policy_preset_options()}
        if normalized not in valid_names:
            normalized = DEFAULT_VENDOR_POLICY_PRESET
        self.app_settings["default_vendor_policy_preset"] = normalized
        session = getattr(self, "session", None)
        if session is not None:
            session.default_vendor_policy_preset = normalized
        self._save_app_settings()

    def _start_update_check(self):
        app_runtime_flow.start_update_check(self, APP_VERSION, is_release_version, threading.Thread)

    def _check_for_updates_worker(self):
        app_runtime_flow.check_for_updates_worker(
            self,
            app_version=APP_VERSION,
            fetch_latest_release=fetch_latest_github_release,
            is_newer_version=is_newer_version,
            url_error_types=(urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError, json.JSONDecodeError),
        )

    def _prompt_for_update(self, release):
        app_runtime_flow.prompt_for_update(
            self,
            release,
            app_version=APP_VERSION,
            releases_page_url=GITHUB_RELEASES_PAGE_URL,
        )

    def _load_gif_frames(self):
        """Load animated GIF frames for the loading overlay."""
        if self._loading_frames_loaded:
            return
        self._loading_frames_loaded = True
        self._loading_frames = loading_flow.load_gif_frames(
            LOADING_GIF_FILE,
            target_size=(200, 200),
            max_frames=MAX_LOADING_GIF_FRAMES,
            has_pil=HAS_PIL,
            image_module=Image if HAS_PIL else None,
            image_tk_module=ImageTk if HAS_PIL else None,
            tk_module=tk,
        )

    def _load_corner_gif_frames(self):
        """Load small animated GIF frames for the notebook corner."""
        if self._corner_loading_frames_loaded:
            return
        self._corner_loading_frames_loaded = True
        self._corner_loading_frames = loading_flow.load_gif_frames(
            LOADING_GIF_FILE,
            target_size=CORNER_LOADING_GIF_SIZE,
            max_frames=MAX_LOADING_GIF_FRAMES,
            has_pil=HAS_PIL,
            image_module=Image if HAS_PIL else None,
            image_tk_module=ImageTk if HAS_PIL else None,
            tk_module=tk,
        )

    def _show_corner_loading_gif(self):
        if not self._corner_loading_frames_loaded:
            self._load_corner_gif_frames()
        label = loading_flow.ensure_corner_loading_gif(
            self,
            lambda parent: tk.Label(parent, bg="#1e1e2e", borderwidth=0, highlightthickness=0),
        )
        if label is None or self._corner_loading_after_id is not None:
            return
        self._corner_loading_frame_idx = 0
        self._animate_corner_loading()

    def _show_loading(self, text="Loading..."):
        """Show the dancing cat loading overlay."""
        if self._loading_overlay is not None:
            return  # already showing
        if not self._loading_frames_loaded:
            self._load_gif_frames()
        self._start_loading_audio()
        if not self._loading_frames:
            # No gif available — just show text
            self._loading_overlay = tk.Frame(self.root, bg="#1e1e2e")
            self._loading_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
            tk.Label(self._loading_overlay, text=text, font=("Segoe UI", 14),
                     fg="#c9a0dc", bg="#1e1e2e").place(relx=0.5, rely=0.5, anchor="center")
            self.root.update()
            return

        self._loading_overlay = tk.Frame(self.root, bg="#1e1e2e")
        self._loading_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._loading_img_label = tk.Label(self._loading_overlay, bg="#1e1e2e")
        self._loading_img_label.place(relx=0.5, rely=0.40, anchor="center")

        self._loading_text_label = tk.Label(
            self._loading_overlay, text=text, font=("Segoe UI", 13, "bold"),
            fg="#c9a0dc", bg="#1e1e2e"
        )
        self._loading_text_label.place(relx=0.5, rely=0.70, anchor="center")

        self._loading_frame_idx = 0
        self._animate_loading()

    def _start_loading_audio(self):
        """Start looping Nyan Cat loading audio if available."""
        loading_flow.start_loading_audio(
            has_winsound=HAS_WINSOUND,
            loading_wav_file=LOADING_WAV_FILE,
            winsound_module=winsound if HAS_WINSOUND else None,
        )

    def _stop_loading_audio(self):
        """Stop loading audio playback."""
        loading_flow.stop_loading_audio(
            has_winsound=HAS_WINSOUND,
            winsound_module=winsound if HAS_WINSOUND else None,
        )

    def _animate_loading(self):
        """Cycle through gif frames at native speed."""
        loading_flow.animate_loading(self)

    def _animate_corner_loading(self):
        """Cycle through gif frames in the notebook corner."""
        loading_flow.animate_corner_loading(self)

    def _autosize_dialog(self, dlg, min_w=420, min_h=280, max_w_ratio=0.9, max_h_ratio=0.9):
        """Size a popup to its content while keeping it inside the screen."""
        loading_flow.autosize_dialog(dlg, min_w=min_w, min_h=min_h, max_w_ratio=max_w_ratio, max_h_ratio=max_h_ratio)

    def _hide_loading(self):
        """Remove the loading overlay."""
        loading_flow.hide_loading(self)

    def _run_with_loading(self, text, func, *args, min_seconds=5):
        """Show loading overlay with animation, run func in a thread, then hide."""
        return loading_flow.run_with_loading(self, text, func, *args, min_seconds=min_seconds)

    # ── Tab 1: Load Files ────────────────────────────────────────────────

    def _build_tab_load(self):
        ui_load.build_load_tab(self)

    def _browse_folder(self):
        path = filedialog.askdirectory(title="Select Folder Containing X4 Report CSVs")
        if path:
            self.var_scan_dir.set(path)

    def _scan_folder(self):
        folder = self.var_scan_dir.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showinfo("No Folder", "Please select a valid folder first.")
            return

        found = parsers.scan_directory_for_reports(folder)

        var_map = {
            "sales": self.var_sales_path,
            "minmax": self.var_minmax_path,
            "onhand": self.var_onhand_path,
            "po": self.var_po_path,
            "susp": self.var_susp_path,
            "packsize": self.var_packsize_path,
        }

        report_names = {
            "sales": "Part Sales & Receipts",
            "minmax": "On Hand Min/Max Sales",
            "onhand": "On Hand Report",
            "po": "POs by PG",
            "susp": "Suspended Items",
            "packsize": "Order Multiples",
        }

        populated = []
        for rtype, filepath in found.items():
            if rtype in var_map:
                var_map[rtype].set(filepath)
                populated.append(report_names.get(rtype, rtype))

        if populated:
            self.lbl_scan_status.config(
                text=f"✓  Found {len(populated)} report(s): {', '.join(populated)}"
            )
        else:
            self.lbl_scan_status.config(text="No X4 report CSVs detected in that folder.")

    def _browse(self, which):
        path = filedialog.askopenfilename(
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if path:
            var_map = {
                "sales": self.var_sales_path,
                "po": self.var_po_path,
                "susp": self.var_susp_path,
                "minmax": self.var_minmax_path,
                "onhand": self.var_onhand_path,
                "packsize": self.var_packsize_path,
            }
            if which in var_map:
                var_map[which].set(path)

    def _do_load(self):
        sales_path = self.var_sales_path.get().strip()
        if not sales_path:
            messagebox.showerror("Missing File", "The Part Sales & Receipts CSV is required.")
            return

        # Gather paths before showing loading screen
        paths = {
            "sales": sales_path,
            "po": self.var_po_path.get().strip(),
            "susp": self.var_susp_path.get().strip(),
            "onhand": self.var_onhand_path.get().strip(),
            "minmax": self.var_minmax_path.get().strip(),
            "packsize": self.var_packsize_path.get().strip(),
        }

        try:
            result = self._run_with_loading("Loading files...", self._parse_all_files, paths)
        except Exception as e:
            messagebox.showerror("Parse Error", f"Failed to parse Part Sales CSV:\n{e}")
            return

        if result is None:
            return

        load_flow.apply_load_result(self.session, result)
        self.sales_span_days = result.get("sales_span_days")

        if not self.sales_items:
            messagebox.showwarning("No Data", "No items found in the Part Sales CSV. Check the file format.")
            return

        # Show warnings from parsing
        for warning in result.get("warnings", []):
            messagebox.showwarning(warning[0], warning[1])
        if self.pack_size_conflicts:
            sample = ", ".join(sorted(self.pack_size_conflicts)[:12])
            extra = "..." if len(self.pack_size_conflicts) > 12 else ""
            messagebox.showwarning(
                "Pack Fallback Conflict",
                (
                    f"{len(self.pack_size_conflicts)} item code(s) have conflicting order multiples across line codes.\n"
                    "Item-code-level pack fallback was skipped for these items.\n\n"
                    f"Examples: {sample}{extra}"
                ),
            )
            # Add summary row for export.
            self.startup_warning_rows.append({
                "warning_type": "Pack Fallback Conflict",
                "severity": "warning",
                "line_code": "",
                "item_code": "",
                "description": "",
                "reference_date": "",
                "qty": "",
                "po_reference": "",
                "details": (
                    f"{len(self.pack_size_conflicts)} conflicting item code(s). "
                    f"Examples: {sample}{extra}"
                ),
            })

        if hasattr(self, "btn_export_startup_warnings"):
            if self.startup_warning_rows:
                self.btn_export_startup_warnings.config(state="normal")
            else:
                self.btn_export_startup_warnings.config(state="disabled")

        # Prompt immediately after startup warnings so export is not missed.
        if self.startup_warning_rows:
            do_export = messagebox.askyesno(
                "Export Startup Warnings?",
                (
                    f"{len(self.startup_warning_rows)} startup warning row(s) are available.\n\n"
                    "These warnings do not block the run, but they may weaken suggestions or deserve a quick X4 check.\n\n"
                    "Do you want to export the Startup Warnings CSV now?"
                ),
            )
            if do_export:
                self._export_startup_warnings_csv()

        # Status summary
        status_parts = [f"{len(self.sales_items)} items loaded"]
        if self.po_items:
            status_parts.append(f"{len(self.po_items)} open PO lines")
        if self.suspended_set:
            status_parts.append(f"{len(self.suspended_set)} suspended items")
        if self.inventory_lookup:
            status_parts.append(f"{len(self.inventory_lookup)} inventory records")
        if self.pack_size_lookup:
            status_parts.append(f"{len(self.pack_size_lookup)} order multiples")
        if self.pack_size_by_item:
            status_parts.append(f"{len(self.pack_size_by_item)} item-level pack fallbacks")
        status_parts.append(f"{len(self.all_line_codes)} line codes found")
        self.lbl_load_status.config(text="✓  " + "  ·  ".join(status_parts))

        # Populate exclude tab and enable it
        self._populate_exclude_tab()
        self.notebook.tab(1, state="normal")
        self.notebook.select(1)

    def _export_startup_warnings_csv(self):
        """Export startup warning details (old POs and other load-time warnings) to CSV."""
        if not self.startup_warning_rows:
            messagebox.showinfo("No Warnings", "No startup warning details are available to export.")
            return

        default_name = f"Startup_Warnings_{datetime.now().strftime('%Y%m%d')}.csv"
        path = filedialog.asksaveasfilename(
            title="Save Startup Warnings CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
        )
        if not path:
            return

        def _text_code(val):
            """Force Excel to treat codes as text for consistent alignment."""
            txt = str(val).strip() if val is not None else ""
            return f"'{txt}" if txt else ""

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Warning Type",
                    "Severity",
                    "Line Code",
                    "Item Code",
                    "Description",
                    "Reference Date",
                    "Qty",
                    "PO Reference",
                    "Details",
                ])
                for row in self.startup_warning_rows:
                    writer.writerow([
                        row.get("warning_type", ""),
                        row.get("severity", ""),
                        _text_code(row.get("line_code", "")),
                        _text_code(row.get("item_code", "")),
                        row.get("description", ""),
                        row.get("reference_date", ""),
                        row.get("qty", ""),
                        row.get("po_reference", ""),
                        row.get("details", ""),
                    ])
            messagebox.showinfo("Saved", f"Startup warnings exported to:\n{os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export startup warnings CSV:\n{e}")

    @staticmethod
    def _parse_all_files(paths):
        """Parse all CSV files (runs in background thread)."""
        return load_flow.parse_all_files(
            paths,
            old_po_warning_days=OLD_PO_WARNING_DAYS,
            short_sales_window_days=SHORT_SALES_WINDOW_DAYS,
        )

    # ── Tab 2: Exclude Line Codes ────────────────────────────────────────

    def _build_tab_exclude(self):
        ui_filters.build_exclude_tab(self)

    def _populate_exclude_tab(self):
        ui_filters.populate_exclude_tab(self)

    def _toggle_all_lc(self, state):
        for var in self.lc_vars.values():
            var.set(state)
        self._update_lc_count()

    def _update_lc_count(self):
        ui_filters.update_lc_count(self)

    def _do_exclude(self):
        self.excluded_line_codes = {lc for lc, var in self.lc_vars.items() if not var.get()}

        # If there are suspended items, populate and show the customer tab
        if self.suspended_items:
            self._populate_customer_tab()
            self.notebook.tab(2, state="normal")
            self.notebook.select(2)
        else:
            # No suspended items — skip customer filter, go straight to assign
            self._proceed_to_assign()

    def _resolve_pack_size(self, key):
        """Resolve pack size with exact-key and item-code fallback."""
        pack = self.pack_size_lookup.get(key)
        if pack:
            return pack
        generic = self.pack_size_lookup.get(("", key[1]))
        if generic:
            return generic
        return self.pack_size_by_item.get(key[1])

    def _get_x4_pack_size(self, key):
        """Return the original X4 order multiple from the loaded source files."""
        pack = self.pack_size_source_lookup.get(key)
        if pack:
            return pack
        generic = self.pack_size_source_lookup.get(("", key[1]))
        if generic:
            return generic
        return None

    def _default_vendor_for_key(self, key):
        """Use the X4 supplier as the initial vendor when available."""
        return reorder_flow.default_vendor_for_key(self, key)

    def _get_suspense_carry_qty(self, key):
        return persistent_state_flow.get_suspense_carry_qty(self, key)

    def _persist_suspense_carry(self):
        return persistent_state_flow.persist_suspense_carry(self, write_debug)

    def _proceed_to_assign(self):
        """Apply all filters, merge suspended items, and move to bulk vendor assignment."""
        self._show_loading("Crunching numbers...")
        self.root.update()
        try:
            days = self.var_lookback_days.get()
        except Exception:
            days = 14

        try:
            has_items = assignment_flow.prepare_assignment_session(
                self.session,
                excluded_line_codes=self.excluded_line_codes,
                excluded_customers=self.excluded_customers,
                dup_whitelist=self.dup_whitelist,
                ignored_keys=self.ignored_item_keys,
                lookback_days=days,
                order_history_path=self._data_path("order_history"),
                vendor_codes_path=self._data_path("vendor_codes"),
                known_vendors=KNOWN_VENDORS,
                get_suspense_carry_qty=self._get_suspense_carry_qty,
                default_vendor_for_key=self._default_vendor_for_key,
                resolve_pack_size=self._resolve_pack_size,
                suggest_min_max=self._suggest_min_max,
                get_cycle_weeks=self._get_cycle_weeks,
                get_rule_key=get_rule_key,
                default_vendor_policy_preset=self._get_default_vendor_policy_preset(),
            )
        except Exception as exc:
            self._hide_loading()
            messagebox.showerror("Vendor Assignment Error", f"Could not prepare vendor assignment:\n{exc}")
            return

        if not has_items:
            self._hide_loading()
            messagebox.showwarning(
                "No Items",
                "After applying filters, no items remain to order.",
            )
            return


        # ── Enrich items with ordering logic ──
        # Normalize demand_signal to the current reorder cycle period.
        # assignment_flow sets demand_signal as a raw total over the full sales
        # export window; divide it down to one cycle's worth so that e.g.
        # 40 bearings sold over a year on a weekly cycle suggests ~1, not 40.
        reorder_flow.normalize_items_to_cycle(self)

        self._loaded_vendor_codes = list(self.vendor_codes_used)
        self.last_removed_bulk_items = []
        try:
            self._refresh_vendor_inputs()
            self._populate_bulk_tree()
            self.notebook.tab(3, state="normal")
            self.notebook.select(3)
        except Exception as exc:
            self._hide_loading()
            messagebox.showerror("Vendor Assignment Error", f"Could not open vendor assignment:\n{exc}")
            return

        # Populate and go to bulk assign
        self._hide_loading()

    # ── Tab 3: Customer Exclusion ─────────────────────────────────────────

    def _build_tab_customers(self):
        ui_filters.build_customer_tab(self)

    def _populate_customer_tab(self):
        ui_filters.populate_customer_tab(self)

    def _toggle_all_cust(self, state):
        for var in self.cust_vars.values():
            var.set(state)
        self._update_cust_count()

    def _update_cust_count(self):
        ui_filters.update_cust_count(self)

    def _do_customer_exclude(self):
        self.excluded_customers = {code for code, var in self.cust_vars.items() if not var.get()}
        self._proceed_to_assign()

    # ── Tab 4: Bulk Vendor Assignment ────────────────────────────────────

    def _build_tab_bulk_assign(self):
        ui_bulk.build_bulk_tab(self, BULK_EDITABLE_COLS)

    def _populate_bulk_tree(self):
        ui_bulk.populate_bulk_tree(self)

    def _get_cycle_weeks(self):
        """Return the number of weeks for the selected reorder cycle."""
        return reorder_flow.get_cycle_weeks(self)

    def _suggest_min_max(self, key):
        """
        Calculate suggested min/max based on 12-month sales and reorder cycle.
        Min = usage during 1 cycle (reorder point)
        Max = usage during 2 cycles (stock-up target)
        Returns (sug_min, sug_max) or (None, None) if history is too sparse.
        """
        return reorder_flow.suggest_min_max(self, key, MIN_ANNUAL_SALES_FOR_SUGGESTIONS)

    def _find_filtered_item(self, key):
        return ui_bulk.find_filtered_item(self, key)

    @staticmethod
    def _normalize_vendor_code(value):
        return persistent_state_flow.normalize_vendor_code(value)

    def _save_vendor_codes(self):
        persistent_state_flow.save_vendor_codes(self)

    def _save_order_rules(self):
        persistent_state_flow.save_order_rules(self)

    def _save_vendor_policies(self):
        persistent_state_flow.save_vendor_policies(self)

    def _save_duplicate_whitelist(self):
        persistent_state_flow.save_duplicate_whitelist(self)

    def _save_ignored_item_keys(self):
        persistent_state_flow.save_ignored_item_keys(self)

    def _refresh_vendor_inputs(self):
        ui_state_flow.refresh_vendor_inputs(self)

    def _remember_vendor_code(self, vendor):
        return persistent_state_flow.remember_vendor_code(self, vendor)

    def _rename_vendor_code(self, old_vendor, new_vendor):
        normalized = persistent_state_flow.rename_vendor_code(self, old_vendor, new_vendor)
        if normalized:
            self._update_bulk_summary()
            self._update_review_summary()
        return normalized

    def _remove_vendor_code(self, vendor):
        ui_state_flow.remove_vendor_code(self, vendor)

    def _open_vendor_manager(self):
        ui_vendor_manager.open_vendor_manager(self)

    def _get_effective_order_qty(self, item):
        return item_workflow.get_effective_order_qty(item)

    def _set_effective_order_qty(self, item, qty, *, manual_override=False):
        item_workflow.set_effective_order_qty(item, qty, manual_override=manual_override)

    @staticmethod
    def _clear_manual_override(item):
        item_workflow.clear_manual_override(item)

    def _annotate_release_decisions(self):
        session = getattr(self, "session", self)
        if hasattr(session, "default_vendor_policy_preset"):
            session.default_vendor_policy_preset = self._get_default_vendor_policy_preset()
        shipping_flow.annotate_release_decisions(session)

    def _recalculate_item(self, item, annotate_release=True):
        session = getattr(self, "session", self)
        item["reorder_cycle_weeks"] = self._get_cycle_weeks()
        item_workflow.recalculate_item_from_session(item, session, self._suggest_min_max, get_rule_key)
        if annotate_release:
            shipping_flow.annotate_release_decisions(session)
        return item

    def _effective_order_rule(self, item, rule):
        return item_workflow.effective_order_rule(item, rule, self.inventory_lookup)

    def _sync_review_item_to_filtered(self, item):
        session = getattr(self, "session", self)
        result = item_workflow.sync_review_item_to_filtered_from_session(
            item,
            session,
            self._suggest_min_max,
            get_rule_key,
        )
        shipping_flow.annotate_release_decisions(session)
        return result

    def _bulk_row_values(self, item):
        return ui_bulk.bulk_row_values(self, item)

    def _bulk_row_id(self, item):
        return ui_bulk.bulk_row_id(item)

    def _resolve_bulk_row_id(self, row_id):
        return ui_bulk.resolve_bulk_row_id(self, row_id)

    def _refresh_suggestions(self):
        """Recalculate suggestions when the reorder cycle changes."""
        reorder_flow.refresh_suggestions(self)

    def _refresh_recent_orders(self):
        """Reload recent orders when lookback days changes."""
        reorder_flow.refresh_recent_orders(self)

    def _update_bulk_summary(self):
        ui_bulk.update_bulk_summary(self)

    def _refresh_bulk_view_after_edit(self, row_ids, changed_cols=None):
        return ui_bulk.refresh_bulk_view_after_edit(self, row_ids, changed_cols=changed_cols)

    def _apply_bulk_filter(self):
        ui_bulk.apply_bulk_filter(self)

    def _sort_bulk_tree(self, col):
        ui_bulk.sort_bulk_tree(self, col)

    def _bulk_vendor_autocomplete(self, event):
        ui_assignment_actions.bulk_vendor_autocomplete(self, event)

    def _bulk_apply_selected(self):
        ui_assignment_actions.bulk_apply_selected(self)

    def _bulk_apply_visible(self):
        ui_assignment_actions.bulk_apply_visible(self)

    def _not_needed_reason(self, item):
        return ui_bulk_dialogs.not_needed_reason(self, item, MAX_EXCEED_ABS_BUFFER)

    def _bulk_remove_not_needed_visible(self, include_assigned=None):
        self._bulk_remove_not_needed(scope="screen", include_assigned=include_assigned)

    def _bulk_remove_not_needed_filtered(self, include_assigned=None):
        """Review and remove currently filtered rows that appear unnecessary to order."""
        self._bulk_remove_not_needed(scope="filtered", include_assigned=include_assigned)

    def _bulk_remove_not_needed(self, scope="screen", include_assigned=None):
        ui_bulk_dialogs.bulk_remove_not_needed(
            self,
            scope,
            MAX_EXCEED_ABS_BUFFER,
            include_assigned=include_assigned,
        )

    def _undo_last_bulk_removal(self):
        ui_assignment_actions.undo_last_bulk_removal(self)

    def _update_bulk_cell_status(self):
        ui_state_flow.update_bulk_cell_status(self)

    def _update_bulk_sheet_status(self):
        self._update_bulk_cell_status()

    def _bulk_copy_selection(self, event=None):
        return bulk_sheet_actions_flow.bulk_copy_selection(self)

    def _bulk_paste_selection(self, event=None):
        return bulk_sheet_actions_flow.bulk_paste_selection(self)

    def _bulk_select_current_row(self, event=None):
        return bulk_sheet_actions_flow.bulk_select_current_row(self)

    def _bulk_select_current_column(self, event=None):
        return bulk_sheet_actions_flow.bulk_select_current_column(self)

    def _capture_bulk_history_state(self, *, capture_spec=None):
        return session_state_flow.capture_bulk_history_state(self, capture_spec=capture_spec)

    def _finalize_bulk_history_action(self, label, before_state, *, coalesce_key=None, capture_spec=None):
        return session_state_flow.finalize_bulk_history_action(
            self,
            label,
            before_state,
            MAX_BULK_HISTORY,
            coalesce_key=coalesce_key,
            capture_spec=capture_spec,
        )

    def _restore_bulk_history_state(self, state, *, capture_spec=None):
        session_state_flow.restore_bulk_history_state(self, state, capture_spec=capture_spec)

    def _bulk_undo(self, event=None):
        if not self.bulk_undo_stack:
            return "break" if event is not None else None
        entry = self.bulk_undo_stack.pop()
        capture_spec = copy.deepcopy(entry.get("_capture_spec"))
        try:
            current_state = self._capture_bulk_history_state(capture_spec=capture_spec)
        except TypeError:
            current_state = self._capture_bulk_history_state()
        try:
            self._restore_bulk_history_state(entry["before"], capture_spec=capture_spec)
        except TypeError:
            self._restore_bulk_history_state(entry["before"])
        self.bulk_redo_stack.append({
            "label": entry.get("label", ""),
            "before": copy.deepcopy(entry["before"]),
            "after": current_state,
            "_coalesce_key": copy.deepcopy(entry.get("_coalesce_key")),
            "_capture_spec": capture_spec,
        })
        return "break" if event is not None else None

    def _bulk_redo(self, event=None):
        if not self.bulk_redo_stack:
            return "break" if event is not None else None
        entry = self.bulk_redo_stack.pop()
        capture_spec = copy.deepcopy(entry.get("_capture_spec"))
        try:
            current_state = self._capture_bulk_history_state(capture_spec=capture_spec)
        except TypeError:
            current_state = self._capture_bulk_history_state()
        try:
            self._restore_bulk_history_state(entry["after"], capture_spec=capture_spec)
        except TypeError:
            self._restore_bulk_history_state(entry["after"])
        self.bulk_undo_stack.append({
            "label": entry.get("label", ""),
            "before": current_state,
            "after": copy.deepcopy(entry["after"]),
            "_coalesce_key": copy.deepcopy(entry.get("_coalesce_key")),
            "_capture_spec": capture_spec,
        })
        return "break" if event is not None else None

    def _bulk_select_all(self, event=None):
        return bulk_sheet_actions_flow.bulk_select_all(self)

    def _bulk_clear_selection(self, event=None):
        return bulk_sheet_actions_flow.bulk_clear_selection(self)

    def _bulk_fill_selection_with_current_value(self, event=None, *, alias="fill"):
        return bulk_sheet_actions_flow.bulk_fill_selection_with_current_value(
            self,
            BULK_EDITABLE_COLS,
            write_debug,
            event,
            alias=alias,
        )

    def _bulk_fill_down_selection(self, event=None):
        return self._bulk_fill_selection_with_current_value(event, alias="fill_down")

    def _bulk_fill_right_selection(self, event=None):
        return self._bulk_fill_selection_with_current_value(event, alias="fill_right")

    def _bulk_apply_current_value_to_selection(self, event=None):
        return self._bulk_fill_selection_with_current_value(event, alias="ctrl_enter")

    def _bulk_move_next_editable_cell(self, event=None):
        return bulk_sheet_actions_flow.bulk_move_next_editable_cell(self)

    def _bulk_move_prev_editable_cell(self, event=None):
        return bulk_sheet_actions_flow.bulk_move_prev_editable_cell(self)

    def _bulk_extend_selection_up(self, event=None):
        return bulk_sheet_actions_flow.bulk_extend_selection_up(self)

    def _bulk_extend_selection_down(self, event=None):
        return bulk_sheet_actions_flow.bulk_extend_selection_down(self)

    def _bulk_extend_selection_left(self, event=None):
        return bulk_sheet_actions_flow.bulk_extend_selection_left(self)

    def _bulk_extend_selection_right(self, event=None):
        return bulk_sheet_actions_flow.bulk_extend_selection_right(self)

    def _bulk_jump_home(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_home(self)

    def _bulk_jump_end(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_end(self)

    def _bulk_jump_ctrl_left(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_ctrl_left(self)

    def _bulk_jump_ctrl_right(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_ctrl_right(self)

    def _bulk_jump_ctrl_up(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_ctrl_up(self)

    def _bulk_jump_ctrl_down(self, event=None):
        return bulk_sheet_actions_flow.bulk_jump_ctrl_down(self)

    def _show_bulk_shortcuts(self):
        messagebox.showinfo("Bulk Shortcuts", BULK_SHORTCUTS_TEXT)

    def _bulk_begin_edit(self, event=None):
        return bulk_sheet_actions_flow.bulk_begin_edit(
            self,
            BULK_EDITABLE_COLS,
            simpledialog.askstring,
            write_debug,
            event,
        )

    def _bulk_begin_edit_from_menu(self):
        write_debug(
            "bulk_begin_edit.menu_command",
            right_click_context=repr(getattr(self, "_right_click_bulk_context", None)),
        )
        return self._bulk_begin_edit()

    def _bulk_fit_columns(self):
        if self.bulk_sheet:
            self.bulk_sheet.fit_columns_to_window()

    def _bulk_remove_selected_rows(self, event=None):
        return bulk_sheet_actions_flow.bulk_remove_selected_rows(
            self,
            copy.deepcopy,
            messagebox.askyesno,
            event,
        )

    def _bulk_fill_selected_cells(self):
        bulk_sheet_actions_flow.bulk_fill_selected_cells(
            self,
            BULK_EDITABLE_COLS,
            simpledialog.askstring,
            messagebox.showinfo,
        )

    def _bulk_clear_selected_cells(self):
        bulk_sheet_actions_flow.bulk_clear_selected_cells(
            self,
            BULK_EDITABLE_COLS,
            messagebox.showinfo,
        )

    def _bulk_delete_selected(self, event=None):
        return bulk_sheet_actions_flow.bulk_delete_selected(self, event)

    def _bulk_apply_editor_value(self, row_id, col_name, raw):
        bulk_edit_flow.apply_editor_value(self, row_id, col_name, raw, BULK_EDITABLE_COLS, get_rule_key, write_debug)

    def _apply_bulk_filter(self):
        ui_bulk.apply_bulk_filter(self)
        sample = ""
        if self.filtered_items:
            item = self.filtered_items[0]
            sample = f"{item.get('line_code', '')}:{item.get('item_code', '')}"
        write_debug("bulk_filter.applied", total=len(self.filtered_items), sample=sample)

    def _open_buy_rule_editor(self, idx):
        bulk_context_flow.open_buy_rule_editor(self, idx, write_debug)

    def _view_item_details(self):
        ui_bulk_dialogs.view_item_details(self)

    def _edit_buy_rule_from_bulk(self):
        ui_bulk_dialogs.edit_buy_rule_from_bulk(self)

    def _resolve_review_from_bulk(self):
        ui_bulk_dialogs.resolve_review_from_bulk(self)

    def _dismiss_duplicate_from_bulk(self):
        ui_bulk_dialogs.dismiss_duplicate_from_bulk(self)

    def _dismiss_duplicate(self, item_code):
        """Add an item code to the persistent duplicate whitelist."""
        bulk_context_flow.dismiss_duplicate(self, item_code)

    @staticmethod
    def _ignore_key(line_code, item_code):
        return session_state_flow.ignore_key(line_code, item_code)

    def _ignore_items_by_keys(self, ignore_keys):
        return session_state_flow.ignore_items_by_keys(self, ignore_keys)

    def _ignore_from_bulk(self):
        bulk_context_flow.ignore_from_bulk(self, messagebox.askyesno, messagebox.showinfo)

    def _go_to_individual(self):
        ui_assignment_actions.go_to_individual(self)

    def _finish_bulk(self):
        """Check stock warnings, then collect assigned items and go to review."""
        if not self._check_stock_warnings():
            return
        self._finish_bulk_final()

    def _check_stock_warnings(self):
        return ui_bulk_dialogs.check_stock_warnings(self)

    def _finish_bulk_final(self):
        ui_bulk_dialogs.finish_bulk_final(self)

    # Tab 5: Individual Vendor Assignment
    # ── Tab 5: Individual Vendor Assignment ───────────────────────────────

    def _build_tab_individual_assign(self):
        ui_individual.build_individual_tab(self)

    def _vendor_autocomplete(self, event):
        ui_assignment_actions.vendor_autocomplete(self, event)

    def _populate_assign_item(self):
        ui_individual.populate_assign_item(self)

    def _dismiss_dup_from_individual(self):
        ui_assignment_actions.dismiss_dup_from_individual(self)

    def _assign_current(self):
        ui_assignment_actions.assign_current(self)

    def _assign_skip(self):
        ui_assignment_actions.assign_skip(self)

    def _assign_back(self):
        ui_assignment_actions.assign_back(self)

    def _finish_assign(self):
        ui_assignment_actions.finish_assign(self)

    # ── Tab 6: Review & Export ───────────────────────────────────────────

    def _build_tab_review(self):
        ui_review.build_review_tab(self)

    def _build_tab_help(self):
        ui_help.build_help_tab(self)

    def _populate_review_tab(self):
        ui_review.populate_review_tab(self)

    def _review_row_values(self, item):
        return ui_review.review_row_values(item)

    def _update_review_summary(self):
        ui_review.update_review_summary(self)

    def _apply_review_filter(self):
        ui_review.apply_review_filter(self)

    def _sort_tree(self, col):
        ui_review.sort_tree(self, col)

    def _review_editor_widget(self, col_name):
        return review_flow.review_editor_widget(self, col_name)

    def _review_editor_value(self, row_id, col_name):
        return review_flow.review_editor_value(self, row_id, col_name)

    def _review_refresh_editor_row(self, row_id):
        review_flow.review_refresh_editor_row(self, row_id)

    def _review_apply_editor_value(self, row_id, col_name, raw):
        review_flow.review_apply_editor_value(self, row_id, col_name, raw, get_rule_key)

    def _on_review_tree_click(self, event):
        """Remember the clicked editable review column for keyboard editing."""
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        col = self.tree.identify_column(event.x)
        if not col:
            return
        col_names = ("vendor", "line_code", "item_code", "description", "order_qty", "status", "why", "pack_size")
        col_idx = int(col.replace("#", "")) - 1
        if 0 <= col_idx < len(col_names):
            self.review_grid_editor.remember_col(col_names[col_idx])

    def _on_review_tree_keyboard_edit(self, event=None):
        """Open the review-tree editor from the keyboard."""
        return self.review_grid_editor.keyboard_edit()

    def _on_review_tree_horizontal_nav(self, event):
        """Change the active editable review column with arrow keys."""
        return self.review_grid_editor.horizontal_nav(event.keysym)

    def _on_tree_double_click(self, event):
        """Edit order_qty or vendor on double-click."""
        region = self.tree.identify("region", event.x, event.y)
        if region == "separator":
            self._autosize_review_column_from_x(event.x)
            return "break"
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return

        col_idx = int(col.replace("#", "")) - 1
        col_names = ("vendor", "line_code", "item_code", "description", "order_qty", "status", "why", "pack_size")
        if col_idx >= len(col_names):
            return
        col_name = col_names[col_idx]

        self.review_grid_editor.open_editor(row_id, col_name)

    def _autosize_review_column_from_x(self, x):
        col = self.tree.identify_column(x)
        if not col:
            return
        col_idx = int(col.replace("#", "")) - 1
        columns = getattr(self, "review_tree_columns", ())
        labels = getattr(self, "review_tree_labels", {})
        if 0 <= col_idx < len(columns):
            col_name = columns[col_idx]
            max_width = 560 if col_name in ("description", "why") else 220
            self.review_grid_editor.autosize_column(col_name, labels.get(col_name, col_name), max_width=max_width)

    def _delete_selected(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("Confirm Delete", f"Remove {len(selected)} item(s) from the PO?"):
            return
        # Remove from assigned_items (highest index first to avoid shifting)
        indices = sorted([int(s) for s in selected], reverse=True)
        for idx in indices:
            if idx < len(self.assigned_items):
                self.assigned_items.pop(idx)
        self._populate_review_tab()

    def _back_to_assign(self):
        self.notebook.select(3)

    def _do_export(self):
        export_flow.do_export(
            self,
            export_vendor_po,
            self._data_path("order_history"),
            self._data_path("sessions"),
        )

    def _build_maintenance_report(self):
        """
        Build a list of X4 data maintenance items from source/session/suggested state.
        """
        candidates = maintenance_flow.build_maintenance_candidates(
            self.session,
            suggest_min_max=self._suggest_min_max,
            get_x4_pack_size=self._get_x4_pack_size,
        )
        return build_maintenance_report(candidates)

    def _build_session_snapshot(self, output_dir, created_files, maintenance_issues):
        return export_flow.build_session_snapshot(self, output_dir, created_files, maintenance_issues)

    def _export_maintenance_csv(self, issues, output_dir):
        return export_flow.export_maintenance_csv(issues, output_dir)

    def _show_maintenance_report(self, output_dir, issues=None):
        if issues is None:
            issues = self._build_maintenance_report()
        ui_review.show_maintenance_report(self, output_dir, issues)

    def _export_vendor_po(self, vendor, items, output_dir):
        return export_vendor_po(vendor, items, output_dir)


# ─── Entry Point ─────────────────────────────────────────────────────────────

def apply_dark_theme(root):
    """Apply a dark mode theme with purple accents."""
    # Color palette
    BG = "#1e1e2e"           # main background
    BG_LIGHT = "#2a2a3d"     # slightly lighter panels
    BG_WIDGET = "#333348"    # entry/combo backgrounds
    FG = "#e0e0e8"           # main text
    FG_DIM = "#9090a8"       # secondary text
    PURPLE = "#b48ead"       # primary accent
    PURPLE_BRIGHT = "#c9a0dc" # hover / highlight
    PURPLE_DARK = "#7c5e8a"  # pressed / selection
    BORDER = "#444460"       # subtle borders
    RED = "#f07070"          # warnings
    TREE_BG = "#252538"      # treeview background
    TREE_ALT = "#2c2c42"     # alternate row
    TREE_SEL = "#5b4670"     # selected row

    root.configure(bg=BG)
    root.option_add("*TCombobox*Listbox.background", BG_WIDGET)
    root.option_add("*TCombobox*Listbox.foreground", FG)
    root.option_add("*TCombobox*Listbox.selectBackground", PURPLE_DARK)
    root.option_add("*TCombobox*Listbox.selectForeground", FG)

    style = ttk.Style()
    style.theme_use("clam")

    # ── Global defaults ──
    style.configure(".", background=BG, foreground=FG, fieldbackground=BG_WIDGET,
                     bordercolor=BORDER, troughcolor=BG_LIGHT, font=("Segoe UI", 10),
                     insertcolor=FG, selectbackground=PURPLE_DARK, selectforeground=FG)

    # ── Frames & Labels ──
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=FG)
    style.configure("TLabelframe", background=BG, foreground=PURPLE, bordercolor=BORDER)
    style.configure("TLabelframe.Label", background=BG, foreground=PURPLE, font=("Segoe UI", 10, "bold"))

    # ── Notebook ──
    style.configure("TNotebook", background=BG, bordercolor=BORDER)
    style.configure("TNotebook.Tab", background=BG_LIGHT, foreground=FG_DIM,
                     padding=[14, 7], font=("Segoe UI", 10))
    style.map("TNotebook.Tab",
              background=[("selected", BG), ("active", BG_WIDGET)],
              foreground=[("selected", PURPLE_BRIGHT), ("active", FG)])

    # ── Buttons ──
    style.configure("TButton", background=BG_WIDGET, foreground=FG,
                     bordercolor=BORDER, padding=[10, 5], font=("Segoe UI", 10))
    style.map("TButton",
              background=[("active", PURPLE_DARK), ("pressed", PURPLE)],
              foreground=[("active", FG), ("pressed", FG)])

    style.configure("Big.TButton", background=PURPLE_DARK, foreground=FG,
                     font=("Segoe UI", 10, "bold"), padding=[16, 7], bordercolor=PURPLE)
    style.map("Big.TButton",
              background=[("active", PURPLE), ("pressed", PURPLE_BRIGHT)])

    # ── Entries & Combos ──
    style.configure("TEntry", fieldbackground=BG_WIDGET, foreground=FG,
                     insertcolor=FG, bordercolor=BORDER)
    style.map("TEntry", fieldbackground=[("focus", "#3a3a52")])

    style.configure("TCombobox", fieldbackground=BG_WIDGET, foreground=FG,
                     background=BG_WIDGET, arrowcolor=PURPLE, bordercolor=BORDER)
    style.map("TCombobox", fieldbackground=[("focus", "#3a3a52")],
              background=[("active", BG_LIGHT)])

    # ── Checkbuttons ──
    style.configure("TCheckbutton", background=BG, foreground=FG, indicatorcolor=BG_WIDGET)
    style.map("TCheckbutton",
              background=[("active", BG_LIGHT)],
              indicatorcolor=[("selected", PURPLE), ("active", PURPLE_DARK)])

    # ── Progressbar ──
    style.configure("TProgressbar", background=PURPLE, troughcolor=BG_LIGHT, bordercolor=BORDER)

    # ── Scrollbar ──
    style.configure("TScrollbar", background=BG_LIGHT, troughcolor=BG,
                     bordercolor=BORDER, arrowcolor=PURPLE)
    style.map("TScrollbar", background=[("active", PURPLE_DARK)])

    # ── Treeview ──
    style.configure("Treeview", background=TREE_BG, foreground=FG,
                     fieldbackground=TREE_BG, bordercolor=BORDER, font=("Segoe UI", 9))
    style.configure("Treeview.Heading", background=BG_WIDGET, foreground=PURPLE_BRIGHT,
                     bordercolor=BORDER, font=("Segoe UI", 9, "bold"))
    style.map("Treeview",
              background=[("selected", TREE_SEL)],
              foreground=[("selected", FG)])
    style.map("Treeview.Heading",
              background=[("active", PURPLE_DARK)])

    # ── Custom label styles ──
    style.configure("Header.TLabel", font=("Segoe UI", 13, "bold"), foreground=PURPLE_BRIGHT, background=BG)
    style.configure("SubHeader.TLabel", font=("Segoe UI", 10), foreground=FG_DIM, background=BG)
    style.configure("Info.TLabel", font=("Segoe UI", 9), foreground=FG_DIM, background=BG)
    style.configure("Warning.TLabel", font=("Segoe UI", 9, "bold"), foreground=RED, background=BG)
    style.configure("Path.TLabel", font=("Segoe UI", 8, "italic"), foreground="#7a7a95", background=BG)

    return style


def main():
    root = tk.Tk()
    apply_dark_theme(root)

    # Set custom window icon if available
    if os.path.exists(ICON_FILE):
        try:
            root.iconbitmap(ICON_FILE)
        except Exception:
            pass

    app = POBuilderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
